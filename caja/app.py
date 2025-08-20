from flask import Flask, render_template, request, redirect, session, send_file, jsonify, flash, url_for
import os, sqlite3, re, json, time, random, string
from datetime import datetime, timedelta
from pytz import timezone
from zoneinfo import ZoneInfo
from openpyxl import load_workbook
from io import BytesIO
import smtplib, ssl
from email.message import EmailMessage
from cryptography.hazmat.primitives import hashes
from cryptography.hazmat.primitives.asymmetric import padding
from cryptography.hazmat.primitives.serialization import load_pem_private_key
from itsdangerous import URLSafeTimedSerializer
from urllib.parse import urlencode
import time
import json
from datetime import datetime
from zoneinfo import ZoneInfo
from itsdangerous import URLSafeTimedSerializer, BadSignature, SignatureExpired
import ast
from email.mime.text import MIMEText
from email.utils import formataddr

app = Flask(__name__)
app.secret_key = 'clave-secreta-caja'
BASE_DIR = os.path.dirname(__file__)
DB_PATH = os.path.join(BASE_DIR, "inventario.sqlite3")

SMTP_HOST = os.getenv("SMTP_HOST", "smtp.gmail.com")
SMTP_PORT = int(os.getenv("SMTP_PORT", "587"))
SMTP_USER = os.getenv("SMTP_USER", "sistemasccfnld@laboratoriocelular.net")
SMTP_PASS = os.getenv("SMTP_PASS", "qvdhlnigelqevtnm")
SMTP_FROM_NAME = os.getenv("SMTP_FROM_NAME", "Celulares Crédito Fácil")
SSO_SHARED_SECRET = "cambia_esta_clave_32+caracteres"
INV_URL_BASE = "http://100.92.172.109:5001"

TZ = timezone("America/Monterrey")

SUCURSALES = ["Hidalgo", "Colinas", "Voluntad 1", "Reservas", "Villas"]

NOMBRES_REALES = {
    "admin_ajnd_7204": "Alejandrina",
    "admin_blcn_8532": "Blanca",
    "admin_vnss_3189": "Vanessa",
    "admin_dvcd_6497": "David",
    "admin_vctr_9051": "Victoria",
    "consulta_abav_5921": "Abigail Avila",
    "consulta_ftrr_8350": "Fatima Reyes",
    "consulta_cndz_1043": "Carolina Díaz",
    "consulta_abcr_6619": "Abigail Corona",
    "consulta_lzrz_1284": "Lizeth Ruiz",
    "consulta_lsrv_2390": "Leslie",
    "consulta_lrbn_3017": "Lorena",
    "consulta_brsd_5126": "Briseida",
    "consulta_evln_6029": "Evelin",
    "consulta_mnsr_1045": "Monserrath",
    "consulta_dnlv_2096": "Daniel",
    "consulta_ltzb_8431": "Litzy",
    "consulta_bryn_9327": "Brayan",
    "consulta_angl_0187": "Angela",
    "reparto_jsss_7493": "Jesús",
    "consulta_daniel_6786": "Daniel",
    "admin_cndz_1043": "Carolina (admin)",
    "consulta_mrsa_5709": "Marisa",
    "consulta_angs_1789": "Angeles",
    "consulta_brayan_4812": "Brayan",
    "consulta_ajnd_7204": "Alejandrina (consulta)",
    
}

TZ = timezone("America/Monterrey")

DENOM = [1, 2, 5, 10, 20, 50, 100]

OBJETIVOS_DEFAULT = {
    1: 40,
    2: 40,
    5: 40,
    10: 20,
    20: 20,
    50: 10,
    100: 5
}

def sso_serializer():
    return URLSafeTimedSerializer(SSO_SHARED_SECRET, salt="sso-inv-5001")

def fecha_hoy():
    return datetime.now(TZ).strftime("%d-%m-%Y")

def fecha_hora_actual():
    return datetime.now(TZ).strftime("%d-%m-%Y %H:%M:%S")

def validar_contrasena_segura(contrasena):
    return bool(re.match(r'^(?=.*[a-z])(?=.*[A-Z])(?=.*\d)(?=.*[@$!%*?&])[A-Za-z\d@$!%*?&]{6,}$', contrasena))

def init_db():
    with sqlite3.connect(DB_PATH) as conn:
        c = conn.cursor()
        c.execute('''
            CREATE TABLE IF NOT EXISTS usuarios (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                nombre TEXT UNIQUE,
                contrasena TEXT,
                tipo TEXT,
                sucursal TEXT
            )
        ''')
        conn.commit()

def registrar_log(usuario, tipo, accion):
    fecha_hora = fecha_hora_actual()
    origen = "Caja"
    with sqlite3.connect(DB_PATH) as conn:
        cur = conn.cursor()
        cur.execute("""
            INSERT INTO log_actividad (usuario, tipo, accion, fecha, origen)
            VALUES (?, ?, ?, ?, ?)
        """, (usuario, tipo, accion, fecha_hora, origen))
        conn.commit()

def get_caja():
    with sqlite3.connect(DB_PATH) as con:
        cur = con.cursor()
        cur.execute("SELECT denom, cantidad FROM caja")
        return dict(cur.fetchall())

def ya_envio_feria(sucursal, fecha):
    with sqlite3.connect(DB_PATH) as conn:
        cur = conn.cursor()
        cur.execute("SELECT COUNT(*) FROM ciclos WHERE sucursal = ? AND fecha = ?", (sucursal, fecha))
        count = cur.fetchone()[0]
        return count > 0

def obtener_sucursales():
    with sqlite3.connect(DB_PATH) as conn:
        cur = conn.cursor()
        try:
            cur.execute("SELECT DISTINCT sucursal FROM usuarios WHERE sucursal IS NOT NULL AND sucursal<>'' ORDER BY sucursal")
            rows = cur.fetchall()
            if rows:
                return [r[0] for r in rows]
        except:
            pass
    return ["Hidalgo", "Colinas", "Voluntad 1", "Reservas", "Villas"]

def ensure_column(conn, table, col, ddl):
    cur = conn.cursor()
    cur.execute(f"PRAGMA table_info({table})")
    cols = {r[1].lower() for r in cur.fetchall()}
    if col.lower() not in cols:
        cur.execute(f"ALTER TABLE {table} ADD COLUMN {col} {ddl}")
        conn.commit()

def ensure_ventas_desglose_schema(conn):
    cur = conn.cursor()
    cur.execute("""
        CREATE TABLE IF NOT EXISTS ventas_desglose(
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            fecha TEXT,
            sucursal TEXT,
            usuario TEXT,
            efectivo REAL DEFAULT 0,
            tarjeta REAL DEFAULT 0,
            dolares REAL DEFAULT 0,
            tipo_cambio REAL DEFAULT 0,
            referencia TEXT
        )
    """)
    cur.execute("PRAGMA table_info(ventas_desglose)")
    cols = {r[1] for r in cur.fetchall()}
    defs = {
        "efectivo": "REAL DEFAULT 0",
        "tarjeta": "REAL DEFAULT 0",
        "dolares": "REAL DEFAULT 0",
        "tipo_cambio": "REAL DEFAULT 0",
        "referencia": "TEXT"
    }
    for col, ddl in defs.items():
        if col not in cols:
            cur.execute(f"ALTER TABLE ventas_desglose ADD COLUMN {col} {ddl}")
    conn.commit()

def enviar_codigo_verificacion(destino: str, codigo: str):
    msg = EmailMessage()
    msg["Subject"] = "Código de verificación"
    msg["From"] = f"{SMTP_FROM_NAME} <{SMTP_USER}>"
    msg["To"] = destino
    msg.set_content(f"Tu código de verificación es: {codigo}\nExpira en 15 minutos.")
    try:
        with smtplib.SMTP(SMTP_HOST, SMTP_PORT, timeout=25) as s:
            s.starttls(context=ssl.create_default_context())
            s.login(SMTP_USER, SMTP_PASS)
            s.send_message(msg)
        return True, ""
    except Exception as e1:
        try:
            with smtplib.SMTP_SSL(SMTP_HOST, 465, context=ssl.create_default_context(), timeout=25) as s:
                s.login(SMTP_USER, SMTP_PASS)
                s.send_message(msg)
            return True, ""
        except Exception as e2:
            return False, f"{type(e2).__name__}: {e2}"

def ahora_mx():
    return datetime.now(ZoneInfo("America/Monterrey")).strftime("%d-%m-%Y %H:%M:%S")

def crear_pendiente(conn, tipo, referencia="", sucursal="", detalle_dict=None, creado_por=""):
    cur = conn.cursor()
    cur.execute(
        "INSERT INTO pendientes (tipo, referencia, sucursal, fecha, detalle, estado, creado_por) VALUES (?, ?, ?, ?, ?, 'pendiente', ?)",
        (tipo, referencia, sucursal, ahora_mx(), json.dumps(detalle_dict or {}), creado_por)
    )
    conn.commit()
    return cur.lastrowid

def _destino_panel_caja():
    t = session.get("tipo","")
    if t == "admin":
        return "/panel-admin"
    if t == "consulta":
        return "/panel-consulta"
    if t == "reparto":
        return "/panel-reparto"
    return "/"

def ahora_mx():
    return datetime.now(ZoneInfo("America/Monterrey")).strftime("%d-%m-%Y %H:%M:%S")

def send_email(to_list, subject, html):
    if not to_list:
        return False
    msg = MIMEText(html, "html", "utf-8")
    msg["Subject"] = subject
    msg["From"] = formataddr((SMTP_FROM_NAME, SMTP_USER))  
    msg["To"] = ", ".join(to_list)
    context = ssl.create_default_context()
    with smtplib.SMTP(SMTP_HOST, SMTP_PORT, timeout=20) as server:
        server.starttls(context=context)
        if SMTP_USER:
            server.login(SMTP_USER, SMTP_PASS)
        server.sendmail(SMTP_USER, to_list, msg.as_string())
    return True

def get_admin_emails(conn):
    cur = conn.cursor()
    cur.execute("SELECT correo FROM usuarios WHERE tipo='admin' AND correo IS NOT NULL AND correo<>''")
    return [r[0] for r in cur.fetchall()]

def get_emails_by_usernames(conn, nombres):
    if not nombres: return []
    qmarks = ",".join("?" for _ in nombres)
    cur = conn.cursor()
    cur.execute(f"SELECT correo FROM usuarios WHERE nombre IN ({qmarks}) AND correo IS NOT NULL AND correo<>''", tuple(nombres))
    return [r[0] for r in cur.fetchall()]

def crear_pendiente(conn, tipo, referencia="", sucursal="", detalle_dict=None, creado_por="", involucrados=None):
    cur = conn.cursor()
    cur.execute(
        "INSERT INTO pendientes (tipo, referencia, sucursal, fecha, detalle, estado, creado_por) VALUES (?, ?, ?, ?, ?, 'pendiente', ?)",
        (tipo, referencia, sucursal, ahora_mx(), json.dumps(detalle_dict or {}), creado_por)
    )
    pid = cur.lastrowid
    for u in (involucrados or []):
        cur.execute("INSERT INTO pendientes_involucrados (pendiente_id, usuario, estado) VALUES (?, ?, 'pendiente')", (pid, u))
    conn.commit()

    try:
        mails_invol = get_emails_by_usernames(conn, involucrados or [])
        mails_admin = get_admin_emails(conn)
        mails_watch = get_watchers_emails(conn)
        to_list = list({*mails_invol, *mails_admin, *mails_watch})
        asunto = f"[PENDIENTE] {tipo} — {referencia or 's/ ref'}"
        html = f"""
        <div style="font-family:Segoe UI,Arial,sans-serif">
          <h2>Nuevo pendiente</h2>
          <p><b>Tipo:</b> {tipo} | <b>Ref:</b> {referencia or '-'} | <b>Sucursal:</b> {sucursal or '-'}</p>
          <p><b>Creado por:</b> {creado_por or '-'} | <b>Fecha:</b> {ahora_mx()}</p>
          <p><b>Involucrados:</b> {", ".join(involucrados or []) or "-"}</p>
          <pre style="background:#f6f6f6;padding:10px;border-radius:8px">{json.dumps(detalle_dict or {}, ensure_ascii=False, indent=2)}</pre>
          <p><a href="{request.url_root.rstrip('/')}/pendientes" target="_blank">Abrir PENDIENTES</a></p>
        </div>
        """
        if to_list:
            send_email(to_list, asunto, html)
        cur.execute("UPDATE pendientes_involucrados SET ultimo_correo=?, recordatorios=COALESCE(recordatorios,0)+1 WHERE pendiente_id=?", (ahora_mx(), pid))
        conn.commit()
    except Exception:
        pass

    return pid

def _fetch_pendientes(conn, usuario_actual=None, solo_mios=False):
    cur = conn.cursor()
    if solo_mios:
        cur.execute("""
            SELECT p.id, p.tipo, p.referencia, p.sucursal, p.fecha, p.detalle, p.estado,
                   GROUP_CONCAT(pi.usuario || ':' || pi.estado) AS inv_raw,
                   COALESCE((SELECT estado FROM pendientes_involucrados WHERE pendiente_id=p.id AND usuario=?),'') AS mi_estado
            FROM pendientes p
            LEFT JOIN pendientes_involucrados pi ON pi.pendiente_id=p.id
            WHERE EXISTS (SELECT 1 FROM pendientes_involucrados x WHERE x.pendiente_id=p.id AND x.usuario=?)
            GROUP BY p.id
            ORDER BY p.fecha DESC
        """, (usuario_actual, usuario_actual))
    else:
        cur.execute("""
            SELECT p.id, p.tipo, p.referencia, p.sucursal, p.fecha, p.detalle, p.estado,
                   GROUP_CONCAT(pi.usuario || ':' || pi.estado) AS inv_raw,
                   COALESCE((SELECT estado FROM pendientes_involucrados WHERE pendiente_id=p.id AND usuario=?),'') AS mi_estado
            FROM pendientes p
            LEFT JOIN pendientes_involucrados pi ON pi.pendiente_id=p.id
            GROUP BY p.id
            ORDER BY p.fecha DESC
        """, (usuario_actual or "",))
    rows = cur.fetchall()
    lista = []
    for r in rows:
        inv = []
        if r[7]:
            for x in r[7].split(","):
                if ":" in x:
                    u, s = x.split(":", 1)
                    inv.append({"usuario": u, "estado": s})
        try:
            dct = json.loads(r[5] or "{}")
        except:
            dct = {}
        lista.append({
            "id": r[0], "tipo": r[1], "referencia": r[2] or "", "sucursal": r[3] or "",
            "fecha": r[4] or "", "detalle": dct, "estado": r[6],
            "involucrados": inv, "mi_estado": r[8] or ""
        })
    return lista

def usuarios_para_select():
    with sqlite3.connect(DB_PATH) as conn:
        conn.row_factory = sqlite3.Row
        cur = conn.cursor()

        cur.execute("PRAGMA table_info(usuarios)")
        cols = {row["name"] for row in cur.fetchall()}

        if "nombre_real" in cols:
            cur.execute("""
                SELECT nombre AS usuario, nombre_real AS nombre
                FROM usuarios
                ORDER BY nombre_real, nombre
            """)
            rows = [dict(r) for r in cur.fetchall()]
        else:
            cur.execute("SELECT nombre FROM usuarios ORDER BY nombre")
            rows = [{"usuario": r["nombre"],
                     "nombre": NOMBRES_REALES.get(r["nombre"], r["nombre"])}
                    for r in cur.fetchall()]
    return rows

def get_watchers_emails(conn):
    cur = conn.cursor()
    cur.execute("SELECT correo FROM usuarios WHERE recibe_todos_pendientes=1 AND correo IS NOT NULL AND correo<>''")
    return [r[0] for r in cur.fetchall()]

def is_watcher(conn, user):
    cur = conn.cursor()
    cur.execute("SELECT 1 FROM usuarios WHERE nombre=? AND recibe_todos_pendientes=1", (user,))
    return cur.fetchone() is not None

@app.route("/sso-login-caja")
def sso_login_caja():
    token = request.args.get("token","")
    if not token:
        return redirect("/")

    try:
        data = sso_serializer().loads(token, max_age=300)
    except (SignatureExpired, BadSignature):
        return redirect("/")

    session["usuario"] = data.get("usuario","")
    session["tipo"] = data.get("tipo","")
    session["sucursal"] = data.get("sucursal","")
    session["inicio_sesion"] = datetime.now().isoformat()

    return redirect(_destino_panel_caja())

@app.before_request
def forzar_logout_por_turno():
    usuario = session.get("usuario")
    tipo = session.get("tipo")

    if usuario and tipo in ["consulta", "reparto"]:
        ahora = datetime.now()
        inicio_sesion_str = session.get("inicio_sesion")

        if not inicio_sesion_str:
            session["inicio_sesion"] = ahora.isoformat()
        else:
            inicio_sesion = datetime.fromisoformat(inicio_sesion_str)
            if ahora - inicio_sesion > timedelta(hours=4):
                session.clear()
                flash("🔒 Tu sesión ha expirado después del turno de 4 horas. Por favor inicia sesión de nuevo.")
                return redirect("/")

@app.route("/", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        tipo_usuario = request.form["tipo"]
        nombre = request.form["usuario"]
        contraseña = request.form["contrasena"]  
        sucursal = request.form.get("sucursal")

        conn = sqlite3.connect(DB_PATH)
        cur = conn.cursor()
        cur.execute("SELECT id, contraseña, tipo FROM usuarios WHERE nombre = ?", (nombre,))
        usuario = cur.fetchone()
        conn.close()

        if usuario and usuario[1] == contraseña and usuario[2] == tipo_usuario:
            session["usuario"] = nombre
            session["tipo"] = tipo_usuario
            session["sucursal"] = sucursal
            session["inicio_sesion"] = datetime.now().isoformat()

            registrar_log(nombre, tipo_usuario, "Inicio de sesión exitoso")

            if tipo_usuario == "admin":
                return redirect("/panel-admin")
            elif tipo_usuario == "consulta":
                return redirect("/panel-consulta")
            elif tipo_usuario == "reparto":
                return redirect("/panel-reparto")
            else:
                return "Tipo de usuario desconocido", 403
        else:
            return render_template("login.html", error="Credenciales incorrectas", tipo=tipo_usuario)

    return render_template("login.html")

@app.route("/panel-admin")
def panel_admin():
    if "usuario" not in session or session.get("tipo") != "admin":
        return redirect("/")

    bloqueos_pendientes = 0
    with sqlite3.connect(DB_PATH) as conn:
        cur = conn.cursor()
        cur.execute("SELECT COUNT(*) FROM bloqueos_feria WHERE autorizado = 0")
        bloqueos_pendientes = cur.fetchone()[0]

    return render_template("panel-admin.html", sucursales=SUCURSALES, bloqueos_pendientes=bloqueos_pendientes)

@app.route("/panel-consulta")
def panel_consulta():
    if "usuario" not in session or session.get("tipo") != "consulta":
        return redirect("/")

    nombre_usuario = session.get("usuario", "")
    nombre_real = NOMBRES_REALES.get(nombre_usuario, nombre_usuario)

    return render_template(
        "panel-consulta.html",
        sucursal=session.get("sucursal", ""),
        tipo=session.get("tipo", ""),
        nombre_real=nombre_real
    )

@app.route("/crear-pendiente", methods=["GET","POST"])
def crear_pendiente_view():
    if "usuario" not in session or session.get("tipo") not in ("admin","consulta"):
        return redirect("/")

    if request.method == "POST":
        tipo = request.form.get("tipo","otro")
        referencia = request.form.get("referencia","")
        sucursal = request.form.get("sucursal","")
        detalle_txt = request.form.get("detalle","")
        involucrados = request.form.getlist("involucrados")   
        try:
            detalle = json.loads(detalle_txt) if detalle_txt.strip().startswith("{") else {"texto": detalle_txt}
        except Exception:
            detalle = {"texto": detalle_txt}
        with sqlite3.connect(DB_PATH) as conn:
            crear_pendiente(conn, tipo, referencia, sucursal, detalle,
                            session.get("usuario",""), involucrados)
        flash("Pendiente creado")
        return redirect("/pendientes")

    usuarios = usuarios_para_select()
    return render_template(
        "crear_pendiente.html",
        usuarios=usuarios,                    
        sucursales=SUCURSALES,
        sucursal_default=session.get("sucursal","")
    )

@app.route("/pendientes")
def ver_pendientes():
    if "usuario" not in session:
        return redirect("/")
    with sqlite3.connect(DB_PATH) as conn:
        data = _fetch_pendientes(conn, session.get("usuario",""), solo_mios=False)
    return render_template("ver_pendientes.html", pendientes=data, usuario_actual=session.get("usuario",""), soy_admin=(session.get("tipo")=="admin"))

@app.route("/mis-pendientes")
def mis_pendientes():
    if "usuario" not in session:
        return redirect("/")
    with sqlite3.connect(DB_PATH) as conn:
        data = _fetch_pendientes(conn, session.get("usuario",""), solo_mios=True)
    return render_template("ver_pendientes.html", pendientes=data, usuario_actual=session.get("usuario",""), soy_admin=(session.get("tipo")=="admin"))

@app.route("/pendiente/realizado/<int:pid>", methods=["POST"])
def pendiente_realizado(pid):
    if "usuario" not in session:
        return jsonify(success=False)
    usuario = session.get("usuario","")
    with sqlite3.connect(DB_PATH) as conn:
        cur = conn.cursor()
        cur.execute("UPDATE pendientes_involucrados SET estado='realizado' WHERE pendiente_id=? AND usuario=?", (pid, usuario))
        conn.commit()
        cur.execute("SELECT COUNT(*) FROM pendientes_involucrados WHERE pendiente_id=? AND estado='pendiente'", (pid,))
        quedan = int(cur.fetchone()[0] or 0)
        if quedan == 0:
            cur.execute("UPDATE pendientes SET estado='autorizado', autorizado_por=?, fecha_autorizacion=? WHERE id=?", (usuario, ahora_mx(), pid))
            conn.commit()
    return jsonify(success=True, quedan=quedan)

@app.route("/pendientes-resumen-json")
def pendientes_resumen_json():
    if "usuario" not in session:
        return jsonify({"ok": False})
    with sqlite3.connect(DB_PATH) as conn:
        cur = conn.cursor()
        cur.execute("SELECT COUNT(*) FROM pendientes WHERE estado='pendiente'")
        total = int(cur.fetchone()[0] or 0)
        cur.execute("SELECT tipo, COUNT(*) FROM pendientes WHERE estado='pendiente' GROUP BY tipo")
        por_tipo = {t: c for t, c in cur.fetchall()}
    return jsonify({"ok": True, "total": total, "por_tipo": por_tipo})

@app.route("/mis-pendientes-resumen-json")
def mis_pendientes_resumen_json():
    if "usuario" not in session:
        return jsonify({"ok": False})
    u = session.get("usuario","")
    with sqlite3.connect(DB_PATH) as conn:
        cur = conn.cursor()
        cur.execute("""SELECT COUNT(*)
                       FROM pendientes_involucrados pi
                       JOIN pendientes p ON p.id=pi.pendiente_id
                       WHERE pi.usuario=? AND pi.estado='pendiente' AND p.estado='pendiente'""", (u,))
        total = int(cur.fetchone()[0] or 0)
    return jsonify({"ok": True, "total": total})

@app.route("/registro", methods=["GET", "POST"])
def registro():
    if request.method == "POST":
        correo = request.form["correo"]
        nombre = request.form["nombre"]
        contrasena = request.form["contrasena"]
        confirmar = request.form["confirmar"]
        telefono = request.form["telefono"]
        tipo_u = request.form["tipo"]

        if contrasena != confirmar:
            return render_template("registro.html", error="Las contraseñas no coinciden.")

        patron = r"^(?=.*[a-z])(?=.*[A-Z])(?=.*\d)(?=.*[!@#$%^&*()_\-+=?<>.,]).{6,}$"
        if not re.match(patron, contrasena):
            return render_template("registro.html", error="La contraseña no cumple con los requisitos.")

        nombre_base = nombre.split()[0].lower()
        codigo = random.randint(1000, 9999)
        usuario_generado = f"{tipo_u}_{nombre_base}_{codigo}"

        with sqlite3.connect(DB_PATH) as conn:
            cur = conn.cursor()
            cur.execute("SELECT 1 FROM usuarios WHERE nombre = ?", (usuario_generado,))
            while cur.fetchone():
                codigo = random.randint(1000, 9999)
                usuario_generado = f"{tipo_u}_{nombre_base}_{codigo}"
                cur.execute("SELECT 1 FROM usuarios WHERE nombre = ?", (usuario_generado,))

            detalle = {
                "nombre_generado": usuario_generado,
                "contrasena": contrasena,
                "tipo": tipo_u,
                "correo": correo,
                "telefono": telefono,
                "nombre_capturado": nombre
            }
            crear_pendiente(conn, "registro_usuario", referencia=usuario_generado, sucursal="", detalle_dict=detalle, creado_por=usuario_generado)

        return render_template("registro.html", success="Tu solicitud fue enviada. Un administrador debe autorizarla.")
    return render_template("registro.html")

@app.route("/recuperar", methods=["GET", "POST"])
def recuperar():
    if request.method == "POST":
        contacto = request.form.get("contacto", "").strip()
        if not contacto:
            return render_template("recuperar.html", error="Debes ingresar tu correo, teléfono o nombre de usuario.")
        codigo = "".join(random.choices(string.digits, k=6))
        expira = int(time.time()) + 15*60
        with sqlite3.connect(DB_PATH) as conn:
            cur = conn.cursor()
            cur.execute("""
                SELECT nombre, correo, telefono
                FROM usuarios
                WHERE correo = ? OR telefono = ? OR nombre = ?
                LIMIT 1
            """, (contacto, contacto, contacto))
            row = cur.fetchone()
            if not row:
                return render_template("recuperar.html", error="No se encontró ninguna cuenta con ese dato.")
            usuario, correo, telefono = row
            if not correo:
                return render_template("recuperar.html", error="Tu cuenta no tiene un correo registrado.")
            cur.execute("""
                CREATE TABLE IF NOT EXISTS codigos_verificacion(
                    usuario TEXT,
                    codigo TEXT,
                    contacto TEXT,
                    expira INTEGER
                )
            """)
            cur.execute("DELETE FROM codigos_verificacion WHERE usuario = ?", (usuario,))
            cur.execute("""
                INSERT INTO codigos_verificacion(usuario, codigo, contacto, expira)
                VALUES (?, ?, ?, ?)
            """, (usuario, codigo, correo, expira))
            conn.commit()
        session["usuario_recuperacion"] = usuario
        ok, err = enviar_codigo_verificacion(correo, codigo)
        if ok:
            return redirect(url_for("verificar_codigo", usuario=usuario))
        else:
            return render_template("recuperar.html", error=f"Error enviando el correo: {err}")
    return render_template("recuperar.html")

@app.route("/verificar-codigo", methods=["GET", "POST"])
def verificar_codigo():
    if request.method == "POST":
        usuario = request.form.get("usuario","").strip()
        codigo_ingresado = request.form.get("codigo","").strip()
        nueva_contrasena = request.form.get("nueva_contrasena","")
        if not usuario or not codigo_ingresado or not nueva_contrasena:
            return render_template("verificar_codigo.html", usuario=usuario, error="Faltan datos.")
        with sqlite3.connect(DB_PATH) as conn:
            cur = conn.cursor()
            cur.execute("SELECT codigo, expira FROM codigos_verificacion WHERE usuario = ? LIMIT 1", (usuario,))
            row = cur.fetchone()
        if not row:
            return render_template("verificar_codigo.html", usuario=usuario, error="No hay un código activo. Vuelve a solicitarlo.")
        codigo_guardado, expira = row
        if int(time.time()) > int(expira):
            return render_template("verificar_codigo.html", usuario=usuario, error="El código ha expirado.")
        if codigo_ingresado != str(codigo_guardado):
            return render_template("verificar_codigo.html", usuario=usuario, error="Código incorrecto.")
        with sqlite3.connect(DB_PATH) as conn:
            cur = conn.cursor()
            cur.execute("UPDATE usuarios SET contraseña = ? WHERE nombre = ?", (nueva_contrasena, usuario))
            cur.execute("DELETE FROM codigos_verificacion WHERE usuario = ?", (usuario,))
            conn.commit()
        session.pop("usuario_recuperacion", None)
        return redirect("/")
    usuario = request.args.get("usuario") or session.get("usuario_recuperacion","")
    return render_template("verificar_codigo.html", usuario=usuario)

@app.route("/enviar-codigo", methods=["POST"])
def enviar_codigo():
    dato = request.form.get("dato","").strip()
    if not dato:
        return render_template("enviar_codigo.html", error="Ingresa un correo, teléfono o nombre.")
    codigo = "".join(random.choices(string.digits, k=6))
    expira = int(time.time()) + 15*60
    with sqlite3.connect(DB_PATH) as conn:
        cur = conn.cursor()
        cur.execute("""
            SELECT nombre, correo, telefono
            FROM usuarios
            WHERE correo = ? OR telefono = ? OR nombre = ?
            LIMIT 1
        """, (dato, dato, dato))
        row = cur.fetchone()
        if not row:
            return render_template("enviar_codigo.html", error="No se encontró ninguna cuenta con ese dato.")
        usuario, correo, telefono = row
        if not correo:
            return render_template("enviar_codigo.html", error="Tu cuenta no tiene un correo registrado.")
        cur.execute("""
            CREATE TABLE IF NOT EXISTS codigos_verificacion(
                usuario TEXT,
                codigo TEXT,
                contacto TEXT,
                expira INTEGER
            )
        """)
        cur.execute("DELETE FROM codigos_verificacion WHERE usuario = ?", (usuario,))
        cur.execute("""
            INSERT INTO codigos_verificacion(usuario, codigo, contacto, expira)
            VALUES (?, ?, ?, ?)
        """, (usuario, codigo, correo, expira))
        conn.commit()
    session["usuario_recuperacion"] = usuario
    ok, err = enviar_codigo_verificacion(correo, codigo)
    if ok:
        return redirect(url_for("verificar_codigo", usuario=usuario))
    else:
        return render_template("enviar_codigo.html", error=f"No se pudo enviar el correo: {err}")

@app.route("/pendiente/autorizar/<int:pid>", methods=["POST"])
def autorizar_pendiente(pid):
    if "usuario" not in session or session.get("tipo") != "admin":
        return jsonify(success=False)
    admin = session.get("usuario","")
    with sqlite3.connect(DB_PATH) as conn:
        cur = conn.cursor()
        cur.execute("SELECT id, tipo, detalle, estado FROM pendientes WHERE id = ?", (pid,))
        row = cur.fetchone()
        if not row:
            return jsonify(success=False, msg="No existe")
        if row[3] != "pendiente":
            return jsonify(success=False, msg="Ya procesado")
        tipo = row[1]
        detalle = {}
        try:
            detalle = json.loads(row[2] or "{}")
        except:
            detalle = {}

        if tipo == "registro_usuario":
            nombre = detalle.get("nombre_generado","")
            contrasena = detalle.get("contrasena","")
            tipo_u = detalle.get("tipo","consulta")
            correo = detalle.get("correo","")
            telefono = detalle.get("telefono","")
            with conn:
                c2 = conn.cursor()
                c2.execute("SELECT 1 FROM usuarios WHERE nombre = ?", (nombre,))
                if c2.fetchone():
                    pass
                else:
                    c2.execute("INSERT INTO usuarios (nombre, contraseña, tipo, correo, telefono, aprobado) VALUES (?, ?, ?, ?, ?, 1)", (nombre, contrasena, tipo_u, correo, telefono))

        if tipo == "feria":
            pass

        cur.execute("UPDATE pendientes SET estado='autorizado', autorizado_por=?, fecha_autorizacion=? WHERE id=?", (admin, ahora_mx(), pid))
        conn.commit()
    return jsonify(success=True)

@app.route("/pendiente/rechazar/<int:pid>", methods=["POST"])
def rechazar_pendiente(pid):
    if "usuario" not in session or session.get("tipo") != "admin":
        return jsonify(success=False)
    admin = session.get("usuario","")
    with sqlite3.connect(DB_PATH) as conn:
        cur = conn.cursor()
        cur.execute("SELECT estado FROM pendientes WHERE id = ?", (pid,))
        row = cur.fetchone()
        if not row:
            return jsonify(success=False)
        if row[0] != "pendiente":
            return jsonify(success=False)
        cur.execute("UPDATE pendientes SET estado='rechazado', autorizado_por=?, fecha_autorizacion=? WHERE id=?", (admin, ahora_mx(), pid))
        conn.commit()
    return jsonify(success=True)

@app.route('/firmar', methods=['POST'])
def firmar():
    from cryptography.hazmat.primitives import hashes
    from cryptography.hazmat.primitives.asymmetric import padding
    from cryptography.hazmat.primitives.serialization import load_pem_private_key
    import base64

    with open("QZ Tray Demo Cert/private-key.pem", "rb") as f:
        key = load_pem_private_key(f.read(), password=None)

    data = request.data
    firma = key.sign(data, padding.PKCS1v15(), hashes.SHA1())
    return base64.b64encode(firma).decode()

@app.route("/logout")
def logout():
    registrar_log(session.get("usuario", "Desconocido"), session.get("tipo", "Desconocido"), "Cerró sesión")
    session.clear()
    return redirect("/")

@app.route("/ir-inventario")
def ir_inventario():
    if "usuario" not in session:
        return redirect("/login")
    payload = {
        "usuario": session.get("usuario", ""),
        "tipo": session.get("tipo", ""),
        "sucursal": session.get("sucursal", ""),
        "ts": int(time.time())
    }
    token = sso_serializer().dumps(payload)
    q = urlencode({"token": token, "next": "/dashboard"})
    return redirect(f"{INV_URL_BASE}/sso-login?{q}")


@app.route("/limpiar-carrito", methods=["POST"])
def limpiar_carrito():
    session.pop("carrito", None)
    session.pop("total", None)
    return '', 204

@app.route("/log-actividad", methods=["GET"])
def log_actividad():
    if "usuario" not in session or session.get("tipo") != "admin":
        return redirect("/")

    origen = request.args.get("origen", "")
    TZ = pytz.timezone("America/Monterrey")

    with sqlite3.connect(DB_PATH) as conn:
        conn.row_factory = sqlite3.Row
        cur = conn.cursor()

        if origen:
            cur.execute("SELECT usuario, tipo, accion, fecha, origen FROM log_actividad WHERE origen = ? ORDER BY id DESC", (origen,))
        else:
            cur.execute("SELECT usuario, tipo, accion, fecha, origen FROM log_actividad ORDER BY id DESC")
        rows = cur.fetchall()

    logs = []
    for row in rows:
        try:
            fecha = datetime.strptime(row["fecha"], "%d-%m-%Y %H:%M:%S")
            fecha_local = fecha.astimezone(TZ)
            fecha_formateada = fecha_local.strftime("%d-%m-%Y %H:%M:%S")
        except Exception:
            fecha_formateada = row["fecha"]

        logs.append({
            "usuario": row["usuario"],
            "tipo": row["tipo"],
            "accion": row["accion"],
            "fecha": fecha_formateada,
            "origen": row["origen"]
        })

    return render_template("log_actividad.html", logs=logs, origen=origen)

@app.route("/gastos", methods=["GET", "POST"])
def gastos():
    if "usuario" not in session:
        return redirect("/")

    tipo_usuario   = session.get("tipo")
    sucursal_sesion= session.get("sucursal")
    hoy_dt         = datetime.now(TZ)
    hoy_dmy        = hoy_dt.strftime("%d-%m-%Y")
    hoy_iso        = hoy_dt.strftime("%Y-%m-%d")
    sucursales     = ["Hidalgo", "Colinas", "Voluntad 1", "Reservas", "Villas"]

    sucursal_filtro = request.args.get("sucursal", sucursal_sesion)
    fecha_qs_iso    = request.args.get("fecha", hoy_iso)   

    try:
        fecha_filtro_db = datetime.strptime(fecha_qs_iso, "%Y-%m-%d").strftime("%d-%m-%Y")
    except Exception:
        fecha_filtro_db = hoy_dmy
        fecha_qs_iso    = hoy_iso

    if tipo_usuario != "admin":
        sucursal_filtro = sucursal_sesion
        fecha_filtro_db = hoy_dmy
        fecha_qs_iso    = hoy_iso

    if request.method == "POST":
        motivo = request.form["motivo"]
        monto  = float(request.form["monto"])

        if tipo_usuario == "admin":
            fecha_post_iso = request.form.get("fecha", hoy_iso)
            try:
                fecha_post_db = datetime.strptime(fecha_post_iso, "%Y-%m-%d").strftime("%d-%m-%Y")
            except Exception:
                fecha_post_db = hoy_dmy
                fecha_post_iso= hoy_iso
            sucursal_post = request.form.get("sucursal", sucursal_sesion)
        else:
            fecha_post_db  = hoy_dmy
            fecha_post_iso = hoy_iso
            sucursal_post  = sucursal_sesion

        with sqlite3.connect(DB_PATH) as conn:
            cur = conn.cursor()
            cur.execute(
                "INSERT INTO gastos (motivo, monto, fecha, sucursal) VALUES (?, ?, ?, ?)",
                (motivo, monto, fecha_post_db, sucursal_post)
            )
            conn.commit()

        registrar_log(session["usuario"], tipo_usuario,
                      f"Registró un gasto en {sucursal_post} de ${monto:.2f} - {motivo}")

        return redirect(f"/gastos?sucursal={sucursal_filtro}&fecha={fecha_qs_iso}")

    with sqlite3.connect(DB_PATH) as conn:
        conn.row_factory = sqlite3.Row
        cur = conn.cursor()
        cur.execute("""
            SELECT id, motivo, monto, fecha, sucursal
            FROM gastos
            WHERE sucursal = ? AND fecha = ?
            ORDER BY id DESC
        """, (sucursal_filtro, fecha_filtro_db))
        gastos_rows = cur.fetchall()

    return render_template(
        "gastos.html",
        tipo=tipo_usuario,
        sucursales=sucursales,
        gastos=gastos_rows,
        sucursal_filtro=sucursal_filtro,
        fecha_filtro=fecha_qs_iso  
    )

@app.route("/eliminar-gasto", methods=["POST"])
def eliminar_gasto():
    if "usuario" not in session or session.get("tipo") != "admin":
        return redirect("/")
    gasto_id = request.form["id"]
    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()
    cur.execute("DELETE FROM gastos WHERE id = ?", (gasto_id,))
    conn.commit()
    conn.close()
    return redirect("/gastos")

@app.route("/ventas", methods=["GET", "POST"])
def ventas():
    if "usuario" not in session:
        return redirect("/login")

    if "carrito" not in session:
        session["carrito"] = []

    tipo_usuario = session.get("tipo")
    sucursal_sesion = session.get("sucursal", "")
    fecha_hoy_dt = datetime.now(TZ)
    fecha_hoy_str = fecha_hoy_dt.strftime("%d-%m-%Y")

    
    fecha_input = request.args.get("fecha")
    if fecha_input:
        try:
            fecha_filtro = datetime.strptime(fecha_input, "%Y-%m-%d").strftime("%d-%m-%Y")
        except ValueError:
            fecha_filtro = fecha_hoy_str
    else:
        fecha_filtro = fecha_hoy_str

    
    sucursal_filtro = request.args.get("sucursal", sucursal_sesion)

    if request.method == "POST":
        descripcion = request.form.get("descripcion", "")
        tipo = request.form.get("tipo", "")
        concepto = request.form.get("concepto", "")
        imei = request.form.get("imei", "")
        referencia = request.form.get("referencia", "")
        tipo_pago = request.form.get("tipo_pago", "")
        precio = float(request.form.get("precio", "0") or 0)

        session["carrito"].append({
            "descripcion": descripcion,
            "tipo": tipo,
            "concepto": concepto,
            "imei": imei,
            "referencia": referencia,
            "tipo_pago": tipo_pago,
            "precio": precio
        })
        session.modified = True
        return redirect("/ventas")

    total = sum(item["precio"] for item in session.get("carrito", []))

    
    ventas_registradas = []
    with sqlite3.connect(DB_PATH) as conn:
        conn.row_factory = sqlite3.Row
        cur = conn.cursor()

        if tipo_usuario == "admin":
            cur.execute("SELECT * FROM ventas WHERE fecha LIKE ? AND sucursal = ?",
                        (f"{fecha_filtro}%", sucursal_filtro))
        elif tipo_usuario == "consulta":
            cur.execute("SELECT * FROM ventas WHERE fecha LIKE ? AND sucursal = ?",
                        (f"{fecha_hoy_str}%", sucursal_sesion))
        else:
            cur.execute("SELECT * FROM ventas WHERE 1=0")

        ventas_registradas = cur.fetchall()

    
    fecha_input_str = datetime.strptime(fecha_filtro, "%d-%m-%Y").strftime("%Y-%m-%d")

    return render_template("ventas.html",
                           carrito=session["carrito"],
                           total=total,
                           tipo=tipo_usuario,
                           ventas=ventas_registradas,
                           sucursales=SUCURSALES,
                           sucursal_filtro=sucursal_filtro,
                           fecha=fecha_input_str)

@app.route('/eliminar-venta-admin/<int:id>', methods=['POST'])
def eliminar_venta_admin(id):
    if "usuario" not in session or session.get("tipo") != "admin":
        return jsonify(success=False, msg="no-auth")
    try:
        with sqlite3.connect(DB_PATH) as conn:
            conn.row_factory = sqlite3.Row
            cur = conn.cursor()

            cur.execute("""
                SELECT id, fecha, sucursal,
                       IFNULL(precio,0)        AS precio,
                       IFNULL(tipo_pago,'')    AS tipo_pago,
                       IFNULL(referencia,'')   AS referencia
                FROM ventas
                WHERE id = ?
            """, (id,))
            vw = cur.fetchone()
            if not vw:
                return jsonify(success=False, msg="venta-no-encontrada")

            fecha_ts  = (vw["fecha"] or "").strip()
            sucursal  = (vw["sucursal"] or "").strip()
            monto_mxn = float(vw["precio"] or 0.0)
            tipo_pago = (vw["tipo_pago"] or "").strip().lower()
            ref       = (vw["referencia"] or "").strip()
            fecha_dia = fecha_ts.split(" ")[0] if " " in fecha_ts else fecha_ts
            like_fecha = (fecha_dia + "%") if fecha_dia else "%"  

            prefer = {
                "efectivo": ["efectivo", "tarjeta", "dolares"],
                "tarjeta":  ["tarjeta", "efectivo", "dolares"],
                "dolares":  ["dolares", "efectivo", "tarjeta"],
            }.get(tipo_pago, ["efectivo", "tarjeta", "dolares"])

            cur.execute("PRAGMA table_info(ventas_desglose)")
            cols = {r["name"].lower() for r in cur.fetchall()}
            rec_col = "id" if "id" in cols else "rowid"

            def _retira_mxn(mxn, con_ref=True):
                if mxn <= 1e-6:
                    return 0.0
                retirado = 0.0
                extra = ""
                params = [like_fecha, sucursal]
                if con_ref and ref:
                    extra = "AND IFNULL(referencia,'') = ?"
                    params.append(ref)
                while mxn > 1e-6:
                    try:
                        cur.execute(f"""
                            SELECT {rec_col} AS rid,
                                   IFNULL(efectivo,0)    AS efectivo,
                                   IFNULL(tarjeta,0)     AS tarjeta,
                                   IFNULL(dolares,0)     AS dolares,
                                   IFNULL(tipo_cambio,0) AS tipo_cambio
                            FROM ventas_desglose
                            WHERE fecha LIKE ? AND sucursal = ? {extra}
                              AND (IFNULL(efectivo,0) > 0 OR IFNULL(tarjeta,0) > 0 OR IFNULL(dolares,0) > 0)
                            ORDER BY rid DESC
                            LIMIT 1
                        """, params)
                        r = cur.fetchone()
                        if not r:
                            break

                        rid = r["rid"]
                        e   = float(r["efectivo"] or 0.0)
                        t   = float(r["tarjeta"]  or 0.0)
                        d   = float(r["dolares"]  or 0.0)
                        tc  = float(r["tipo_cambio"] or 0.0)

                        disp_mxn = {
                            "efectivo": e,
                            "tarjeta":  t,
                            "dolares":  (d * tc if tc > 0 else 0.0),
                        }
                        orden = prefer + [c for c in ("efectivo","tarjeta","dolares") if c not in prefer]

                        for col in orden:
                            if mxn <= 1e-6:
                                break
                            disponible = disp_mxn[col]
                            if disponible <= 1e-6:
                                continue
                            tomar_mxn = min(mxn, disponible)

                            if col == "dolares":
                                if tc <= 0:
                                    continue
                                tomar_usd = tomar_mxn / tc
                                cur.execute(
                                    f"UPDATE ventas_desglose "
                                    f"SET dolares = CASE WHEN dolares - ? < 0 THEN 0 ELSE dolares - ? END "
                                    f"WHERE {rec_col} = ?",
                                    (tomar_usd, tomar_usd, rid)
                                )
                            elif col == "efectivo":
                                cur.execute(
                                    f"UPDATE ventas_desglose "
                                    f"SET efectivo = CASE WHEN efectivo - ? < 0 THEN 0 ELSE efectivo - ? END "
                                    f"WHERE {rec_col} = ?",
                                    (tomar_mxn, tomar_mxn, rid)
                                )
                            else:
                                cur.execute(
                                    f"UPDATE ventas_desglose "
                                    f"SET tarjeta = CASE WHEN tarjeta - ? < 0 THEN 0 ELSE tarjeta - ? END "
                                    f"WHERE {rec_col} = ?",
                                    (tomar_mxn, tomar_mxn, rid)
                                )

                            mxn      -= tomar_mxn
                            retirado += tomar_mxn

                            cur.execute(
                                f"DELETE FROM ventas_desglose "
                                f"WHERE {rec_col} = ? "
                                f"AND IFNULL(efectivo,0) <= 1e-6 "
                                f"AND IFNULL(tarjeta,0)  <= 1e-6 "
                                f"AND IFNULL(dolares,0)  <= 1e-6",
                                (rid,)
                            )
                    except Exception:
                        break
                return retirado

            retirado = _retira_mxn(monto_mxn, con_ref=True)
            faltante = monto_mxn - retirado
            if faltante > 1e-6:
                _retira_mxn(faltante, con_ref=False)

            cur.execute("DELETE FROM ventas WHERE id = ?", (id,))
            conn.commit()

        return jsonify(success=True)
    except Exception as e:
        print("Error al eliminar venta:", repr(e))
        return jsonify(success=False, msg=str(e))

@app.route("/eliminar-item", methods=["POST"])
def eliminar_item():
    if "usuario" not in session or "carrito" not in session:
        return jsonify(success=False), 403

    data = request.get_json()
    index = int(data.get("index"))

    if 0 <= index < len(session["carrito"]):
        session["carrito"].pop(index)
        session.modified = True

    total = sum(item["precio"] for item in session["carrito"])
    return jsonify(success=True, total=total)

from datetime import datetime
import pytz

TZ = pytz.timezone("America/Monterrey")

def fecha_hoy():
    return datetime.now(TZ).strftime("%d-%m-%Y")

@app.route("/cobrar", methods=["POST"])
def cobrar():
    if "usuario" not in session or "carrito" not in session:
        return redirect("/login")

    carrito = session["carrito"]
    usuario = session["usuario"]
    tipo_usuario = session.get("tipo", "")

    tz = ZoneInfo("America/Monterrey")
    ahora = datetime.now(tz)

    if tipo_usuario == "admin":
        sucursal = (request.form.get("sucursal_manual") or session.get("sucursal", "")).strip()
        fecha_manual = (request.form.get("fecha_manual") or "").strip()
        if fecha_manual:
            try:
                fecha_dt = datetime.strptime(fecha_manual, "%Y-%m-%d").replace(tzinfo=tz)
                fecha_dt = fecha_dt.replace(hour=ahora.hour, minute=ahora.minute, second=ahora.second)
            except ValueError:
                fecha_dt = ahora
        else:
            fecha_dt = ahora
    else:
        sucursal = session.get("sucursal", "")
        fecha_dt = ahora

    fecha = fecha_dt.strftime("%d-%m-%Y %H:%M:%S")
    fecha_dia = fecha_dt.strftime("%d-%m-%Y")

    def safe_float(v):
        if v is None:
            return 0.0
        v = str(v).strip().replace(",", "")
        if v == "":
            return 0.0
        try:
            return float(v)
        except ValueError:
            return 0.0

    efectivo = safe_float(request.form.get("efectivo"))
    tarjeta  = safe_float(request.form.get("tarjeta"))
    dolares  = safe_float(request.form.get("dolares"))
    tipo_cambio = safe_float(request.form.get("dolar"))
    referencia_general = request.form.get("referencia", "").strip()

    def a_centavos(x):
        return int(round(safe_float(x) * 100))

    def ensure_column(conn, table, col, ddl):
        cur = conn.cursor()
        cur.execute(f"PRAGMA table_info({table})")
        cols = {r[1].lower() for r in cur.fetchall()}
        if col.lower() not in cols:
            cur.execute(f"ALTER TABLE {table} ADD COLUMN {col} {ddl}")
            conn.commit()

    with sqlite3.connect(DB_PATH) as conn:
        cur = conn.cursor()
        ensure_ventas_desglose_schema(conn)
        ensure_column(conn, "ventas", "imei", "TEXT")

        cur.execute("SELECT COUNT(*) FROM ventas WHERE fecha LIKE ? AND sucursal = ?", (f"{fecha_dia}%", sucursal))
        ventas_hoy = cur.fetchone()[0]

        if ventas_hoy == 0:
            if not carrito:
                flash("No hay productos en el carrito.")
                return redirect("/ventas")

            descripcion_primera = str(carrito[0].get("descripcion", "")).strip().upper()
            concepto_primera = str(carrito[0].get("concepto", "")).strip().upper()
            es_feria = (descripcion_primera == "FERIA" or concepto_primera == "FERIA")

            cur.execute("""
                CREATE TABLE IF NOT EXISTS bloqueos_feria (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    sucursal TEXT,
                    fecha TEXT,
                    motivo TEXT,
                    autorizado INTEGER DEFAULT 0
                )
            """)

            cur.execute("SELECT autorizado FROM bloqueos_feria WHERE fecha = ? AND sucursal = ? ORDER BY id DESC LIMIT 1",
                        (fecha_dia, sucursal))
            fila_bloqueo = cur.fetchone()
            autorizado_hoy = (fila_bloqueo and int(fila_bloqueo[0]) == 1)

            if not es_feria and not autorizado_hoy:
                cur.execute("""SELECT id FROM bloqueos_feria
                               WHERE fecha = ? AND sucursal = ? AND motivo = ? AND autorizado = 0
                               ORDER BY id DESC LIMIT 1""",
                            (fecha_dia, sucursal, "Primera venta no fue FERIA"))
                ya = cur.fetchone()
                if not ya:
                    cur.execute("INSERT INTO bloqueos_feria (sucursal, fecha, motivo, autorizado) VALUES (?, ?, ?, 0)",
                                (sucursal, fecha_dia, "Primera venta no fue FERIA"))
                    conn.commit()
                return render_template("venta_bloqueada.html",
                                       mensaje="❌ La primera venta del día debe ser FERIA. Espera autorización de un administrador.")

            if es_feria and not autorizado_hoy:
                monto_venta_feria = safe_float(carrito[0].get("precio", 0))
                ayer = (fecha_dt - timedelta(days=1)).strftime("%d-%m-%Y")
                cur.execute("""SELECT COALESCE(SUM(monto), 0) FROM gastos
                               WHERE fecha = ? AND sucursal = ? AND TRIM(UPPER(motivo)) = 'FERIA'""",
                            (ayer, sucursal))
                total_feria_ayer = safe_float(cur.fetchone()[0])

                if a_centavos(monto_venta_feria) != a_centavos(total_feria_ayer):
                    cur.execute("""SELECT id FROM bloqueos_feria
                                   WHERE fecha = ? AND sucursal = ? AND motivo = ? AND autorizado = 0
                                   ORDER BY id DESC LIMIT 1""",
                                (fecha_dia, sucursal, "Feria no coincide con la salida de ayer"))
                    ya = cur.fetchone()
                    if not ya:
                        cur.execute("INSERT INTO bloqueos_feria (sucursal, fecha, motivo, autorizado) VALUES (?, ?, ?, 0)",
                                    (sucursal, fecha_dia, "Feria no coincide con la salida de ayer"))
                        conn.commit()
                    return render_template("venta_bloqueada.html",
                                           mensaje="❌ El monto de la FERIA no coincide con la salida registrada ayer. Espera autorización de un administrador.")
                else:
                    carrito[0]["precio"] = round(total_feria_ayer, 2)

        total_venta = sum(safe_float(item.get("precio", 0)) for item in carrito)
        total_mxn = round(total_venta, 2)
        usd_mxn_entregado = round(dolares * tipo_cambio, 2)

        restante = total_mxn
        usd_usado_mxn = min(usd_mxn_entregado, restante); restante = round(restante - usd_usado_mxn, 2)
        efectivo_usado  = min(efectivo, restante);        restante = round(restante - efectivo_usado, 2)
        tarjeta_usada   = min(tarjeta, restante);         restante = round(restante - tarjeta_usada, 2)
        dolares_usados  = round(usd_usado_mxn / tipo_cambio, 2) if tipo_cambio > 0 else 0.0

        total_pago = round(efectivo + tarjeta + usd_mxn_entregado, 2)
        cambio = round(total_pago - total_mxn, 2)
        if cambio < 0:
            flash(f"❌ Aún faltan ${abs(cambio):.2f} para completar el pago.")
            return redirect("/ventas")

        for item in carrito:
            ref_item = (item.get("referencia") or "").strip()
            cur.execute("""
                INSERT INTO ventas
                (usuario, sucursal, descripcion, tipo, concepto, referencia, imei, precio, tipo_pago, fecha)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                usuario, sucursal,
                item.get("descripcion", ""),
                item.get("tipo", ""),
                item.get("concepto", ""),
                ref_item if ref_item else referencia_general,
                (item.get("imei") or "").strip(),
                safe_float(item.get("precio", 0)),
                item.get("tipo_pago", ""),
                fecha
            ))

            imei = (item.get("imei") or "").strip()
            if imei:
                cur.execute("UPDATE articulos SET estado = 'Vendido' WHERE imei = ?", (imei,))

        cur.execute("""
            INSERT INTO ventas_desglose
            (fecha, sucursal, usuario, efectivo, tarjeta, dolares, tipo_cambio, referencia)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            fecha, sucursal, usuario,
            round(efectivo_usado, 2),
            round(tarjeta_usada, 2),
            round(dolares_usados, 2),
            round(tipo_cambio, 4),
            referencia_general
        ))

        conn.commit()

    NOMBRES_USUARIOS = {
        "Litzi": "Litzi",
        "Brayan": "Brayan",
        "Jesús": "Jesús",
        "consulta_abav_5921": "Abigail Avila",
        "consulta_ftrr_8350": "Fátima Reyes",
        "consulta_cndz_1043": "Carolina Díaz",
        "consulta_abcr_6619": "Abigail Corona",
        "consulta_lzrz_1284": "Lizeth Ruiz",
        "consulta_lsrv_2390": "Leslie Reservas",
        "consulta_alnd_7452": "Alondra",
        "consulta_lrbn_3017": "Lorena",
        "consulta_brsd_5126": "Briseida",
        "consulta_evln_6029": "Evelin",
        "consulta_mnsr_1045": "Monserrath",
        "consulta_angl_7238": "Angeles",
        "consulta_dnlv_2096": "Daniel",
        "consulta_ltzb_8431": "Litzy",
        "consulta_bryn_9327": "Brayan",
        "consulta_arln_6235": "Arlen",
        "consulta_angl_0187": "Angela",
        "consulta_ajnd_7204": "Alejandrina (consulta)",
        "admin_dvcd_6497": "David",
        "admin_vctr_9051": "Victoria",
        "admin_ajnd_7204": "Alejandrina"
    }

    nombre_usuario = NOMBRES_USUARIOS.get(usuario, usuario)
    productos = carrito.copy()
    session.pop("carrito", None)

    session["ultimo_ticket"] = {
        "sucursal": sucursal,
        "fecha": fecha,
        "nombre_usuario": nombre_usuario,
        "productos": productos,
        "total": round(total_venta, 2),
        "total_pagado": round(total_pago, 2),
        "cambio": round(cambio, 2)
    }

    if tipo_usuario == "admin":
        flash("✅ Venta registrada correctamente.")
        return redirect("/ventas")

    return redirect("/abrir-ticket")
@app.route("/ver-bloqueos")
def ver_bloqueos():
    return redirect("/pendientes")

@app.route("/autorizar-bloqueo/<int:id>")
def autorizar_bloqueo(id):
    return redirect("/pendientes")

@app.route("/bloqueos-pendientes-json")
def bloqueos_pendientes_json():
    fecha_hoy = datetime.now(ZoneInfo("America/Monterrey")).strftime("%d-%m-%Y")
    with sqlite3.connect(DB_PATH) as conn:
        cur = conn.cursor()
        cur.execute("SELECT COUNT(*) FROM pendientes WHERE tipo='feria' AND estado='pendiente' AND fecha LIKE ?", (fecha_hoy+"%",))
        total = int(cur.fetchone()[0] or 0)
        cur.execute("SELECT DISTINCT sucursal FROM pendientes WHERE tipo='feria' AND estado='pendiente' AND fecha LIKE ? ORDER BY sucursal", (fecha_hoy+"%",))
        sucursales = [r[0] for r in cur.fetchall()]
    return jsonify({"bloqueos_pendientes": total, "sucursales": sucursales})

@app.route("/abrir-ticket")
def abrir_ticket():
    return """
    <script>
      window.open('/ultimo-ticket', '_blank', 'width=400,height=600');
      window.location.href = '/ventas';
    </script>
    """

@app.route("/ticket")
def ticket():
    if "usuario" not in session:
        return redirect("/login")

    
    fecha = datetime.now(ZoneInfo("America/Monterrey")).strftime("%d/%m/%Y %H:%M %S (Hora local)")

    
    NOMBRES_USUARIOS = {
        "Litzi": "Litzi",
        "Brayan": "Brayan",
        "Jesús": "Jesús",
        "consulta_abav_5921": "Abigail Avila",
        "consulta_ftrr_8350": "Fátima Reyes",
        "consulta_cndz_1043": "Carolina Díaz",
        "consulta_abcr_6619": "Abigail Corona",
        "consulta_lzrz_1284": "Lizeth Ruiz",
        "consulta_lsrv_2390": "Leslie Reservas",
        "consulta_alnd_7452": "Alondra",
        "consulta_lrbn_3017": "Lorena",
        "consulta_brsd_5126": "Briseida",
        "consulta_evln_6029": "Evelin",
        "consulta_mnsr_1045": "Monserrath",
        "consulta_angl_7238": "Angeles",
        "consulta_dnlv_2096": "Daniel",
        "consulta_ltzb_8431": "Litzy",
        "consulta_bryn_9327": "Brayan",
        "consulta_arln_6235": "Arlen",
        "consulta_angl_0187": "Angela",
        "consulta_ajnd_7204": "Alejandrina (consulta)",
    }

    sucursal = session.get("sucursal", "Sucursal Desconocida")
    usuario = session.get("usuario", "Desconocido")
    nombre_usuario = NOMBRES_USUARIOS.get(usuario, usuario)

    
    productos = [
        {
            "descripcion": "iPhone 13 Pro Max 256GB",
            "concepto": "AT&T",
            "referencia": "1234567890",
            "precio": 19999.99
        }
    ]
    total = sum(item["precio"] for item in productos)
    total_pagado = 20000
    cambio = total_pagado - total

    return render_template("ticket.html",
                           sucursal=sucursal,
                           fecha=fecha,
                           nombre_usuario=nombre_usuario,
                           productos=productos,
                           total=total,
                           total_pagado=total_pagado,
                           cambio=cambio)

@app.route("/ultimo-ticket")
def ultimo_ticket():
    if "ultimo_ticket" not in session:
        return "<h3>No hay ticket reciente para imprimir.</h3>"

    datos = session["ultimo_ticket"]

    return render_template("ticket.html",
        sucursal=datos["sucursal"],
        fecha=datos["fecha"],
        nombre_usuario=datos["nombre_usuario"],
        productos=datos["productos"],
        total=datos["total"],
        total_pagado=datos["total_pagado"],
        cambio=datos["cambio"]
    )

@app.route("/ver-ticket/<int:venta_id>")
def ver_ticket(venta_id):
    if "usuario" not in session:
        return redirect("/")

    with sqlite3.connect(DB_PATH) as conn:
        conn.row_factory = sqlite3.Row
        cur = conn.cursor()
        cur.execute("SELECT * FROM ventas WHERE id = ?", (venta_id,))
        venta = cur.fetchone()

    if not venta:
        return "Venta no encontrada", 404

    return render_template("ticket.html",
        sucursal=venta["sucursal"],
        fecha=venta["fecha"],
        producto=venta["descripcion"],
        precio=venta["precio"],
        metodo_pago=venta["tipo_pago"]
    )

@app.route("/fondo")
def ver_fondo_caja():
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("SELECT fecha, sucursal, cantidad FROM fondo_caja ORDER BY fecha DESC, sucursal ASC")
    registros = c.fetchall()
    conn.close()
    return render_template("fondo_caja.html", registros=registros)

@app.route("/cambio", methods=["GET", "POST"])
def cambio():
    if "usuario" not in session:
        return redirect("/login")

    hoy = get_fecha_hoy()

    if request.method == "POST":
        sucursal = request.form["sucursal"]
        fecha = request.form["fecha"]

        with sqlite3.connect(DB_PATH) as con:
            cur = con.cursor()
            cur.execute("SELECT datos FROM ciclos WHERE sucursal = ? AND fecha = ?", (sucursal, fecha))
            registro_existente = cur.fetchone()
            if registro_existente:
                datos = ast.literal_eval(registro_existente[0])
                session["ciclo"] = datos
                session["ya_registrado"] = True
                return redirect("/cambio/confirmar")

            actuales = {d: int(request.form.get(f"actual_{d}", 0)) for d in DENOM}

        objetivos = OBJETIVOS_DEFAULT.copy()
        diferencias = {d: objetivos[d] - actuales[d] for d in DENOM}
        dar = sum(max(0, diferencias[d]) * d for d in DENOM)
        recibir = sum(abs(min(0, diferencias[d])) * d for d in DENOM)

        repuesto = {}
        with sqlite3.connect(DB_PATH) as con:
            cur = con.cursor()
            for d in DENOM:
                delta = diferencias[d]
                if delta > 0:
                    actual_caja = cur.execute("SELECT cantidad FROM caja WHERE denom = ?", (d,)).fetchone()
                    actual_caja = actual_caja[0] if actual_caja else 0
                    repuesto[d] = min(delta, actual_caja)
                else:
                    repuesto[d] = 0

        total_objetivo = sum(objetivos[d] * d for d in DENOM)
        total_actual = sum(actuales[d] * d for d in DENOM)
        total_repuesto = sum(repuesto[d] * d for d in DENOM)

        session["ciclo"] = {
            "sucursal": sucursal,
            "fecha": fecha,
            "actuales": actuales,
            "objetivos": objetivos,
            "diferencias": diferencias,
            "dar": dar,
            "recibir": recibir,
            "repuesto": repuesto,
            "total_objetivo": total_objetivo,
            "total_actual": total_actual,
            "total_repuesto": total_repuesto
        }
        session["ya_registrado"] = False
        return redirect("/cambio/confirmar")

    sucursal = session.get("sucursal", SUCURSALES[0])
    actuales = {d: 0 for d in DENOM}

    return render_template("cambio_index.html",
                           sucursales=SUCURSALES,
                           fecha_actual=hoy,
                           objetivos=OBJETIVOS_DEFAULT,
                           fondo=get_caja(),
                           actuales=actuales)

@app.route("/cambio/confirmar", methods=["GET", "POST"])
def confirmar_cambio():
    if "ciclo" not in session:
        return redirect("/cambio")

    raw_data = session.get("ciclo")
    data = {
        "sucursal": raw_data["sucursal"],
        "fecha": raw_data["fecha"],
        "actuales": {int(k): v for k, v in raw_data["actuales"].items()},
        "objetivos": {int(k): v for k, v in raw_data["objetivos"].items()},
        "diferencias": {int(k): v for k, v in raw_data["diferencias"].items()},
        "repuesto": {int(k): v for k, v in raw_data["repuesto"].items()},
        "dar": raw_data["dar"],
        "recibir": raw_data["recibir"]
    }

    total_objetivo = raw_data.get("total_objetivo", sum(data["objetivos"][d] * d for d in DENOM))
    total_actual = raw_data.get("total_actual", sum(data["actuales"][d] * d for d in DENOM))
    total_repuesto = raw_data.get("total_repuesto", sum(data["repuesto"][d] * d for d in DENOM))

    ya_registrado = session.get("ya_registrado", False)

    if request.method == "POST" and not ya_registrado:
        with sqlite3.connect(DB_PATH) as con:
            cur = con.cursor()
            for d in DENOM:
                delta = data["diferencias"][d]
                if delta > 0:
                    cur.execute("UPDATE caja SET cantidad = cantidad - ? WHERE denom = ?", (data["repuesto"][d], d))
                elif delta < 0:
                    cur.execute("UPDATE caja SET cantidad = cantidad + ? WHERE denom = ?", (abs(delta), d))
            cur.execute("INSERT INTO ciclos (sucursal, fecha, datos) VALUES (?, ?, ?)",
                        (data["sucursal"], data["fecha"], str(data)))
        session.pop("ciclo")
        session.pop("ya_registrado")
        return redirect("/cambio")

    return render_template("cambio_confirmar.html",
                           data=data,
                           denom=DENOM,
                           total_objetivo=total_objetivo,
                           total_actual=total_actual,
                           total_repuesto=total_repuesto,
                           ya_registrado=ya_registrado)

@app.route("/caja-actual", methods=["GET", "POST"])
def caja_actual():
    if "acceso_caja" not in session:
        return redirect("/acceso-caja")

    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()

    if request.method == "POST":
        for denom, cantidad in request.form.items():
            cur.execute("UPDATE caja SET cantidad = ? WHERE denom = ?", (cantidad, denom))
        conn.commit()
        registrar_log(session["usuario"], session["tipo"], "Modificó valores de 'Caja Actual'")

    cur.execute("SELECT denom, cantidad FROM caja ORDER BY CAST(denom AS INTEGER)")
    fondo = cur.fetchall()
    conn.close()

    return render_template("caja_actual.html", fondo=fondo)

@app.route("/acceso-caja", methods=["GET", "POST"])
def acceso_caja():
    if request.method == "POST":
        clave = request.form.get("clave", "")
        if clave == "262293":
            session["acceso_caja"] = True
            return redirect("/caja-actual")
        else:
            return render_template("acceso_caja.html", error="Contraseña incorrecta")
    return render_template("acceso_caja.html")

@app.route("/ciclo-calculo", methods=["POST"])
def ciclo_calculo():
    if "usuario" not in session or session.get("tipo") != "admin":
        return redirect("/login")

    datos = []
    for den in [1, 2, 5, 10, 20, 50, 100]:
        objetivo = int(request.form.get(f"objetivo_{den}", 0))
        actual = int(request.form.get(f"actual_{den}", 0))
        repuesto = int(request.form.get(f"repuesto_{den}", 0))
        total = actual + repuesto
        diferencia = total - objetivo

        datos.append({
            "denominacion": den,
            "objetivo": objetivo,
            "actual": actual,
            "repuesto": repuesto,
            "total": total,
            "diferencia": diferencia
        })

    return render_template("cambio_confirmar.html", datos=datos)

@app.route("/guardar-ciclo", methods=["POST"])
def guardar_ciclo():
    if "usuario" not in session or session.get("tipo") != "admin":
        return redirect("/login")

    json_datos = request.form["json_datos"]
    datos = ast.literal_eval(json_datos)

    sucursal = session.get("sucursal", "Desconocida")
    fecha = datetime.now(TZ).strftime("%d-%m-%Y")

    with sqlite3.connect(DB_PATH) as conn:
        cur = conn.cursor()
        cur.execute("INSERT INTO ciclos (sucursal, fecha, datos) VALUES (?, ?, ?)", (sucursal, fecha, json_datos))

        for d in datos:
            den = d["denominacion"]
            total = d["actual"] + d["repuesto"]
            diferencia = d["diferencia"]

            cur.execute("UPDATE fondo SET cantidad_actual = ?, cantidad_repuesta = ?, monto_dado = monto_dado + ?, monto_recibido = monto_recibido + ? WHERE denom = ?",
                        (total, d["repuesto"], abs(diferencia) if diferencia < 0 else 0, diferencia if diferencia > 0 else 0, den))

        conn.commit()

    return redirect("/panel-admin")

@app.route("/ver-ciclos")
def ver_ciclos():
    if "usuario" not in session:
        return redirect("/login")

    with sqlite3.connect(DB_PATH) as conn:
        conn.row_factory = sqlite3.Row
        cur = conn.cursor()
        cur.execute("SELECT id, sucursal, fecha, datos FROM ciclos ORDER BY date(fecha) DESC, id DESC")
        rows = cur.fetchall()

    ciclos = []
    for row in rows:
        datos_raw = row["datos"]
        datos = {}
        try:
            if isinstance(datos_raw, dict):
                datos = datos_raw
            else:
                try:
                    datos = json.loads(datos_raw)
                except Exception:
                    datos = ast.literal_eval(datos_raw)
        except Exception:
            datos = {}

        objetivos = datos.get("objetivos", {})
        objetivos = {int(k) if isinstance(k, str) else k: int(v or 0) for k, v in objetivos.items()}

        diferencias = datos.get("diferencias", {})
        dif_norm = {}
        for k, v in diferencias.items():
            kk = int(k) if isinstance(k, str) else k
            if isinstance(v, dict):
                dif_norm[kk] = v
            else:
                try:
                    dif_norm[kk] = int(v or 0)
                except Exception:
                    dif_norm[kk] = 0

        actuales = datos.get("actuales") or {}
        if not actuales:
            tmp = {}
            for k, obj in objetivos.items():
                dv = dif_norm.get(k, dif_norm.get(str(k), 0))
                if isinstance(dv, dict):
                    act = int(dv.get("actual", obj) or 0)
                else:
                    act = int(obj + (dv or 0))
                tmp[k] = act
            actuales = tmp
        else:
            actuales = {int(k) if isinstance(k, str) else k: int(v or 0) for k, v in actuales.items()}

        repuesto = datos.get("repuesto", {})
        repuesto = {int(k) if isinstance(k, str) else k: int(v or 0) for k, v in repuesto.items()}

        desglosado = []
        for denom in sorted(objetivos.keys()):
            obj = int(objetivos.get(denom, 0))
            act = int(actuales.get(denom, 0))
            rep = int(repuesto.get(denom, 0))
            dif = dif_norm.get(denom, 0)
            if isinstance(dif, dict):
                dif_val = int(dif.get("diferencia", act - obj))
            else:
                dif_val = int(dif or (act - obj))
            total = act * int(denom)
            desglosado.append({
                "denom": denom,
                "objetivo": obj,
                "actual": act,
                "repuesto": rep,
                "total": total,
                "diferencia": dif_val
            })

        ciclos.append({
            "id": row["id"],
            "sucursal": row["sucursal"],
            "fecha": row["fecha"],
            "total_dado": int(datos.get("dar", datos.get("total_dar", 0)) or 0),
            "total_recibido": int(datos.get("recibir", datos.get("total_recibir", 0)) or 0),
            "desglosado": desglosado
        })

    return render_template("ver_ciclos.html", ciclos=ciclos)


@app.route("/ver-fondo")
def ver_fondo():
    if "usuario" not in session or session.get("tipo") != "admin":
        return redirect("/")

    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()
    cur.execute("SELECT denom, SUM(cantidad) as total FROM caja GROUP BY denom ORDER BY denom ASC")
    datos = cur.fetchall()
    conn.close()

    denominaciones_validas = [1, 2, 5, 10, 20, 50, 100]

    fondo = []
    total_fondo = 0
    for row in datos:
        denom = row["denom"]
        if denom not in denominaciones_validas:
            continue
        cantidad = row["total"] or 0
        valor = denom * cantidad
        total_fondo += valor
        fondo.append({"denominacion": denom, "cantidad": cantidad, "valor": valor})

    return render_template("ver_fondo.html", fondo=fondo, total_fondo=total_fondo)

@app.route("/modificar-ciclo/<int:ciclo_id>", methods=["GET", "POST"])
def modificar_ciclo(ciclo_id):
    if "usuario" not in session or session.get("tipo") != "admin":
        return redirect("/login")

    with sqlite3.connect(DB_PATH) as conn:
        cur = conn.cursor()
        cur.execute("SELECT sucursal, fecha, datos FROM ciclos WHERE id = ?", (ciclo_id,))
        row = cur.fetchone()

        if not row:
            return "Ciclo no encontrado"

        sucursal, fecha, datos_str = row
        datos = eval(datos_str)

        if request.method == "POST":
            nuevos_datos = {}
            for den in DENOM:
                nuevos_datos[den] = {
                    "objetivo": int(request.form.get(f"objetivo_{den}", 0)),
                    "actual": int(request.form.get(f"actual_{den}", 0)),
                    "repuesto": int(request.form.get(f"repuesto_{den}", 0))
                }

            cur.execute("UPDATE ciclos SET datos = ? WHERE id = ?", (str(nuevos_datos), ciclo_id))
            conn.commit()
            return redirect("/ver-ciclos")

    return render_template("modificar_ciclo.html", ciclo_id=ciclo_id, sucursal=sucursal, fecha=fecha, datos=datos)

@app.route("/faltantes-json")
def faltantes_json():
    import pytz
    from datetime import datetime
    import sqlite3

    sucursales = ["Hidalgo", "Colinas", "Voluntad 1", "Reservas", "Villas"]
    TZ = pytz.timezone("America/Monterrey")
    hoy = datetime.now(TZ).strftime("%Y-%m-%d")

    with sqlite3.connect(DB_PATH) as conn:
        cur = conn.cursor()
        cur.execute("SELECT DISTINCT sucursal FROM ciclos WHERE fecha = ?", (hoy,))
        registradas = [r[0] for r in cur.fetchall()]

    faltan = [s for s in sucursales if s not in registradas]
    return jsonify({
        "faltan": faltan,
        "todas_registradas": len(faltan) == 0
    })

@app.route("/feria", methods=["GET", "POST"])
def feria():
    if "usuario" not in session or session.get("tipo") not in ["admin", "consulta"]:
        return redirect("/")

    hoy = datetime.now(TZ).strftime("%d-%m-%Y")
    sucursal = session.get("sucursal", "Villas")

    if request.method == "POST":
        actuales = {int(k.split("_")[1]): int(v) for k, v in request.form.items() if k.startswith("actual_")}
        objetivos = OBJETIVOS_DEFAULT.copy()

        reponer = {}
        entregar = {}
        total_objetivo = total_actual = total_reponer = total_entregar = 0

        conn = sqlite3.connect(DB_PATH)
        cur = conn.cursor()

        for denom in DENOM:
            objetivo = objetivos.get(denom, 0)
            actual = actuales.get(denom, 0)
            valor_objetivo = objetivo * denom
            valor_actual = actual * denom

            total_objetivo += valor_objetivo
            total_actual += valor_actual

            if actual < objetivo:
                cantidad = objetivo - actual
                reponer[denom] = cantidad
                total_reponer += cantidad * denom

                sugerido = min(cantidad, objetivo)  
                cur.execute("UPDATE fondo SET cantidad_actual = cantidad_actual - ? WHERE denominacion = ?", (sugerido, denom))
            elif actual > objetivo:
                cantidad = actual - objetivo
                entregar[denom] = cantidad
                total_entregar += cantidad * denom

                cur.execute("UPDATE fondo SET cantidad_actual = cantidad_actual + ? WHERE denominacion = ?", (cantidad, denom))

        conn.commit()
        conn.close()

        session["feria_datos"] = {
            "sucursal": sucursal,
            "fecha": hoy,
            "actuales": actuales,
            "reponer": reponer,
            "entregar": entregar,
            "total_objetivo": total_objetivo,
            "total_actual": total_actual,
            "total_reponer": total_reponer,
            "total_entregar": total_entregar
        }

        return redirect("/confirmar-feria")

    actuales = {d: 0 for d in DENOM}
    return render_template("feria.html",
                           objetivos=OBJETIVOS_DEFAULT,
                           actuales=actuales,
                           denom=DENOM,
                           sucursal=sucursal,
                           fecha=hoy)

@app.route("/confirmar-feria", methods=["GET", "POST"])
def confirmar_feria():
    if "usuario" not in session or session.get("tipo") not in ["admin", "consulta"]:
        return redirect("/")

    if request.method == "GET":
        return "No puedes acceder directamente a esta página.", 405

    sucursal = session.get("sucursal", "")
    fecha = datetime.now(TZ).strftime("%d-%m-%Y")

    denominaciones = [1, 2, 5, 10, 20, 50, 100]
    cantidades_objetivo = {1: 40, 2: 40, 5: 40, 10: 20, 20: 20, 50: 10, 100: 5}
    cantidades_actual = {}
    cantidades_reponer = {}
    diferencias = {}

    total_objetivo = 0
    total_actual = 0
    total_entregar = 0
    total_reponer = 0
    total_sobrante = 0

    for denom in denominaciones:
        actual = int(request.form.get(f"actual_{denom}", "0") or "0")
        reponer = int(request.form.get(f"reponer_{denom}", "0") or "0")
        objetivo = cantidades_objetivo[denom]
        diferencia = actual - objetivo
        valor_diferencia = diferencia * denom

        if diferencia > 0:
            total_sobrante += valor_diferencia

        cantidades_actual[denom] = actual
        cantidades_reponer[denom] = reponer
        diferencias[denom] = {
            "objetivo": objetivo,
            "actual": actual,
            "diferencia": diferencia,
            "valor_diferencia": valor_diferencia
        }

        total_objetivo += objetivo * denom
        total_actual += actual * denom
        total_entregar += reponer * denom
        
        if actual < objetivo:
            total_reponer += (objetivo - actual) * denom
        
    data = {
        "sucursal": sucursal,
        "fecha": fecha,
        "diferencias": diferencias,
        "repuesto": cantidades_reponer,
        "total_objetivo": total_objetivo,
        "total_actual": total_actual,
        "total_entregar": total_entregar,
        "total_reponer": total_reponer,
        "total_sobrante": total_sobrante
    }

    return render_template("confirmar-feria.html", data=data, denom=denominaciones)

@app.route("/guardar-feria", methods=["POST"])
def guardar_feria():
    if "usuario" not in session:
        return redirect("/login")

    sucursal = session.get("sucursal")
    tipo = session.get("tipo")
    fecha = datetime.now(TZ).strftime("%Y-%m-%d")

    if ya_envio_feria(sucursal, fecha):
        return "Ya se envió la feria hoy."

    objetivos = {1: 40, 2: 40, 5: 40, 10: 20, 20: 20, 50: 10, 100: 5}
    actuales = {}
    diferencias = {}
    repuesto = {}
    total_dar = total_recibir = 0

    for val in objetivos.keys():
        actual_str = request.form.get(f"actual_{val}", "").strip()
        try:
            actual = int(actual_str) if actual_str else 0
        except ValueError:
            actual = 0

        actuales[val] = actual
        dif = actual - objetivos[val]
        diferencias[val] = dif

        if dif < 0:
            rep = abs(dif)
            repuesto[val] = rep
            total_dar += val * rep
        else:
            repuesto[val] = 0
            total_recibir += val * dif

    with sqlite3.connect(DB_PATH) as conn:
        cur = conn.cursor()

        for val in objetivos:
            cur.execute("SELECT cantidad FROM caja WHERE denom = ?", (val,))
            row = cur.fetchone()
            actual_caja = row[0] if row else 0

            nuevo_valor = actual_caja
            if diferencias[val] < 0:
                nuevo_valor -= abs(diferencias[val])  
            elif diferencias[val] > 0:
                nuevo_valor += abs(diferencias[val])  

            cur.execute("UPDATE caja SET cantidad = ? WHERE denom = ?", (nuevo_valor, val))

        datos = {
            "sucursal": sucursal,
            "fecha": fecha,
            "actuales": actuales,
            "objetivos": objetivos,
            "diferencias": diferencias,
            "repuesto": repuesto,
            "dar": total_dar,
            "recibir": total_recibir
        }

        cur.execute("INSERT INTO ciclos (sucursal, fecha, datos) VALUES (?, ?, ?)", (sucursal, fecha, json.dumps(datos)))
        conn.commit()

    if tipo == "admin":
        return redirect("/panel-admin")
    else:
        return redirect("/panel-consulta")

@app.route("/eliminar-ciclo/<int:ciclo_id>", methods=["GET", "POST"])
def eliminar_ciclo(ciclo_id=None):
    if "usuario" not in session or session.get("tipo") != "admin":
        return redirect("/")

    if request.method == "POST":
        _id = ciclo_id or request.form.get("id")
    else:
        _id = ciclo_id

    try:
        _id = int(_id)
    except (TypeError, ValueError):
        return "ID no válido", 400

    suc = request.values.get("sucursal", "")
    fec = request.values.get("fecha", "")

    with sqlite3.connect(DB_PATH) as conn:
        cur = conn.cursor()
        cur.execute("DELETE FROM ciclos WHERE id = ?", (_id,))
        conn.commit()

    return redirect(f"/ver-ciclos?sucursal={suc}&fecha={fec}" if (suc or fec) else "/ver-ciclos")

@app.route("/feria-admin", methods=["GET","POST"])
def feria_admin():
    if "usuario" not in session or session.get("tipo") != "admin":
        return redirect("/")

    sucursales = obtener_sucursales()
    sucursal_sel = request.values.get("sucursal") or (session.get("sucursal") if session.get("sucursal") in sucursales else sucursales[0])
    fecha_sel = request.values.get("fecha") or datetime.now(TZ).strftime("%Y-%m-%d")

    try:
        fecha_fmt = datetime.strptime(fecha_sel, "%Y-%m-%d").strftime("%d-%m-%Y")
    except ValueError:
        fecha_fmt = datetime.strptime(fecha_sel, "%d-%m-%Y").strftime("%d-%m-%Y")

    if request.method == "POST" and request.form.get("fase") == "calcular":
        denominaciones = [1,2,5,10,20,50,100]
        objetivos = {1:40,2:40,5:40,10:20,20:20,50:10,100:5}
        cantidades_actual = {}
        cantidades_reponer = {}
        diferencias = {}
        total_objetivo = 0
        total_actual = 0
        total_entregar = 0
        total_reponer = 0
        total_sobrante = 0

        for denom in denominaciones:
            actual = int(request.form.get(f"actual_{denom}", "0") or "0")
            rep = int(request.form.get(f"reponer_{denom}", "0") or "0")
            objetivo = objetivos[denom]
            dif = actual - objetivo
            val_dif = dif * denom
            if dif > 0:
                total_sobrante += val_dif
            cantidades_actual[denom] = actual
            cantidades_reponer[denom] = rep
            diferencias[denom] = {"objetivo":objetivo,"actual":actual,"diferencia":dif,"valor_diferencia":val_dif}
            total_objetivo += objetivo * denom
            total_actual += actual * denom
            total_entregar += rep * denom
            if actual < objetivo:
                total_reponer += (objetivo - actual) * denom

        data = {
            "sucursal": sucursal_sel,
            "fecha": fecha_sel,
            "diferencias": diferencias,
            "repuesto": cantidades_reponer,
            "total_objetivo": total_objetivo,
            "total_actual": total_actual,
            "total_entregar": total_entregar,
            "total_reponer": total_reponer,
            "total_sobrante": total_sobrante,
            "modo_admin": True
        }
        return render_template("confirmar-feria.html", data=data, denom=denominaciones)

    actuales = {d:0 for d in [1,2,5,10,20,50,100]}
    return render_template("feria.html",
                           objetivos={1:40,2:40,5:40,10:20,20:20,50:10,100:5},
                           actuales=actuales,
                           denom=[1,2,5,10,20,50,100],
                           sucursal=sucursal_sel,
                           fecha=fecha_fmt,
                           modo_admin=True,
                           sucursales=sucursales)

@app.route("/guardar-feria-admin", methods=["POST"])
def guardar_feria_admin():
    if "usuario" not in session or session.get("tipo") != "admin":
        return redirect("/login")

    def _norm_fecha(s):
        s = (s or "").strip()
        if not s:
            return datetime.now(TZ).strftime("%Y-%m-%d")
        for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%Y/%m/%d"):
            try:
                return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
            except Exception:
                pass
        try:
            return datetime.fromisoformat(s).strftime("%Y-%m-%d")
        except Exception:
            return datetime.now(TZ).strftime("%Y-%m-%d")

    sucursal = request.form.get("sucursal") or session.get("sucursal")
    fecha = _norm_fecha(request.form.get("fecha"))

    if ya_envio_feria(sucursal, fecha):
        return "Ya se envió la feria para esa fecha."

    objetivos = {1:40,2:40,5:40,10:20,20:20,50:10,100:5}
    diferencias = {}
    repuesto = {}
    total_dar = 0
    total_recibir = 0

    for val in objetivos.keys():
        actual_str = (request.form.get(f"actual_{val}", "") or "").strip()
        try:
            actual = int(actual_str) if actual_str else 0
        except ValueError:
            actual = 0
        dif = actual - objetivos[val]
        diferencias[val] = dif
        if dif < 0:
            rep = -dif
            repuesto[val] = rep
            total_dar += val * rep
        else:
            repuesto[val] = 0
            total_recibir += val * dif

    with sqlite3.connect(DB_PATH) as conn:
        cur = conn.cursor()
        for val, dif in diferencias.items():
            cur.execute("SELECT cantidad FROM caja WHERE denom = ?", (val,))
            row = cur.fetchone()
            actual_caja = row[0] if row else 0
            nuevo = actual_caja + (dif if dif > 0 else -abs(dif))
            cur.execute("UPDATE caja SET cantidad = ? WHERE denom = ?", (nuevo, val))
        datos = {
            "sucursal": sucursal,
            "fecha": fecha,
            "objetivos": objetivos,
            "diferencias": diferencias,
            "repuesto": repuesto,
            "dar": total_dar,
            "recibir": total_recibir
        }
        cur.execute("INSERT INTO ciclos (sucursal, fecha, datos) VALUES (?, ?, ?)", (sucursal, fecha, json.dumps(datos)))
        conn.commit()

    return redirect("/panel-admin")

@app.route("/nota", methods=["GET", "POST"])
def nota():
    if "usuario" not in session:
        return redirect("/")

    tipo = session.get("tipo", "")
    es_admin = (tipo == "admin")

    def ymd_to_dmy(s):
        if not s: return s
        if len(s) >= 10 and s[4] == "-" and s[7] == "-":
            y, m, d = s[:10].split("-")
            return f"{d}-{m}-{y}"
        return s

    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()

    if es_admin:
        cur.execute("""
            SELECT sucursal FROM ventas
            UNION
            SELECT sucursal FROM notas
        """)
        sucursales_disponibles = sorted({r[0] for r in cur.fetchall() if r[0]})
        if session.get("sucursal") and session["sucursal"] not in sucursales_disponibles:
            sucursales_disponibles.append(session["sucursal"])
    else:
        sucursales_disponibles = [session.get("sucursal", "")]

    hoy_dt = datetime.now(TZ)
    hoy_str = hoy_dt.strftime("%d-%m-%Y")
    ayer_str = (hoy_dt - timedelta(days=1)).strftime("%d-%m-%Y")

    fecha_qs_raw = request.args.get("fecha")
    fecha_qs = ymd_to_dmy(fecha_qs_raw) if fecha_qs_raw else hoy_str
    sucursal_qs = request.args.get("sucursal") or (
        session.get("sucursal", "") if not es_admin
        else (sucursales_disponibles[0] if sucursales_disponibles else "")
    )

    if request.method == "POST":
        texto = (request.form.get("texto") or "").strip()
        if es_admin:
            fecha_post = ymd_to_dmy(request.form.get("fecha_form")) or fecha_qs
            sucursal_post = request.form.get("sucursal_form") or sucursal_qs
        else:
            fecha_post = fecha_qs
            sucursal_post = session.get("sucursal", "")

        if texto and fecha_post and sucursal_post:
            cur.execute(
                "INSERT INTO notas (texto, fecha, sucursal) VALUES (?, ?, ?)",
                (texto, fecha_post, sucursal_post)
            )
            conn.commit()
            registrar_log(session["usuario"], tipo, f"Registró una nota en {sucursal_post} ({fecha_post})")
            conn.close()
            return redirect(f"/nota?fecha={fecha_post}&sucursal={sucursal_post}")

    cur.execute(
        "SELECT id, texto FROM notas WHERE fecha = ? AND sucursal = ? ORDER BY id DESC",
        (fecha_qs, sucursal_qs)
    )
    notas = cur.fetchall()

    cur.execute(
        "SELECT id, texto, fecha, sucursal FROM notas WHERE fecha = ? AND sucursal = ? ORDER BY id DESC LIMIT 50",
        (fecha_qs, sucursal_qs)
    )
    historial = cur.fetchall()

    conn.close()

    return render_template(
        "notas.html",
        notas=notas,
        es_admin=es_admin,
        fecha_actual=fecha_qs,          
        sucursal_actual=sucursal_qs,
        sucursales=sucursales_disponibles,
        hoy_str=hoy_str,
        ayer_str=ayer_str,
        historial=historial
    )

@app.route("/eliminar-nota", methods=["POST"])
def eliminar_nota():
    if "usuario" not in session:
        return redirect("/")

    tipo = session.get("tipo", "")
    if tipo not in ("consulta", "admin"):
        return redirect("/")

    nota_id = request.form.get("id")
    fecha = request.form.get("fecha") or datetime.now(TZ).strftime("%d-%m-%Y")
    sucursal = request.form.get("sucursal") or session.get("sucursal", "")

    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()

    if tipo == "consulta":
        cur.execute("DELETE FROM notas WHERE id = ? AND sucursal = ?", (nota_id, session.get("sucursal", "")))
    else:
        cur.execute("DELETE FROM notas WHERE id = ?", (nota_id,))

    conn.commit()
    conn.close()

    registrar_log(session["usuario"], tipo, f"Eliminó una nota ID {nota_id}")
    return redirect(f"/nota?fecha={fecha}&sucursal={sucursal}")

@app.route("/chat/enviar", methods=["POST"])
def enviar_mensaje():
    data = request.json
    remitente = session.get("usuario", "anónimo")
    destinatario = data.get("destinatario")
    mensaje = data.get("mensaje")
    modulo = data.get("modulo")
    archivo = data.get("archivo", None)  

    if not mensaje or not destinatario:
        return jsonify({"ok": False, "error": "Mensaje o destinatario faltante"})

    fecha = datetime.now(TZ).strftime("%d-%m-%Y %H:%M:%S")

    with sqlite3.connect(DB_PATH) as con:
        cur = con.cursor()
        cur.execute("""
            INSERT INTO chat (fecha, remitente, destinatario, mensaje, modulo, leido, archivo)
            VALUES (?, ?, ?, ?, ?, 0, ?)
        """, (fecha, remitente, destinatario, mensaje, modulo, archivo))

    return jsonify({"ok": True})

@app.route("/chat/recibir")
def recibir_mensajes():
    usuario = session.get("usuario", "")
    tipo = session.get("tipo", "")
    modulo = request.args.get("modulo", "")

    with sqlite3.connect(DB_PATH) as con:
        con.row_factory = sqlite3.Row
        cur = con.cursor()
        cur.execute("""
            SELECT * FROM chat
            WHERE destinatario IN (?, ?, ?, 'Todos')
            AND modulo IN (?, 'General', '')
            ORDER BY id DESC LIMIT 100
        """, (usuario, session.get("sucursal", ""), tipo, modulo))
        mensajes = cur.fetchall()

    return jsonify([dict(row) for row in mensajes])

@app.route("/chat/notificaciones")
def notificaciones_chat():
    usuario = session.get("usuario", "")
    tipo = session.get("tipo", "")
    modulo = request.args.get("modulo", "")

    with sqlite3.connect(DB_PATH) as con:
        cur = con.cursor()
        cur.execute("""
            SELECT COUNT(*) FROM chat
            WHERE destinatario IN (?, ?, ?, 'Todos') AND leido = 0
            AND modulo IN (?, 'General', '')
        """, (usuario, session.get("sucursal", ""), tipo, modulo))
        total = cur.fetchone()[0]

    return jsonify({"nuevos": total})

@app.route("/subir_archivo_chat", methods=["POST"])
def subir_archivo_chat():
    archivo = request.files.get("archivo")
    if not archivo:
        return jsonify({"ok": False, "error": "No hay archivo"})

    nombre = datetime.now().strftime("%d%m%Y%H%M%S") + "_" + archivo.filename
    ruta = os.path.join("static", "chat", nombre)
    archivo.save(ruta)

    return jsonify({"ok": True, "nombre": nombre})

@app.route("/ver-reporte-excel")
def ver_reporte_excel():
    if "usuario" not in session:
        return redirect("/")
    tipo_usuario = session.get("tipo","")
    suc_sesion = session.get("sucursal","")

    with sqlite3.connect(DB_PATH) as conn:
        cur = conn.cursor()
        cur.execute("""
            SELECT DISTINCT sucursal FROM (
                SELECT sucursal FROM ventas
                UNION
                SELECT sucursal FROM gastos
                UNION
                SELECT sucursal FROM fondo_caja
                UNION
                SELECT sucursal FROM dolares
                UNION
                SELECT sucursal FROM anticipos
                UNION
                SELECT sucursal FROM notas
            ) WHERE sucursal IS NOT NULL AND TRIM(sucursal) <> ''
            ORDER BY sucursal
        """)
        sucursales = [r[0] for r in cur.fetchall()]

    if tipo_usuario != "admin" and suc_sesion:
        if suc_sesion not in sucursales:
            sucursales.insert(0, suc_sesion)

    fecha_iso = request.args.get("fecha", fecha_hoy())
    sucursal = request.args.get("sucursal", suc_sesion)

    return render_template(
        "ver_reporte_excel.html",
        sucursal=sucursal,
        sucursales=sucursales,
        fecha=fecha_iso,
        tipo_usuario=tipo_usuario
    )

@app.route("/api/reporte-excel-resumen")
def api_reporte_excel_resumen():
    if "usuario" not in session:
        return jsonify(ok=False, msg="no-auth")

    tipo_usuario = session.get("tipo","")
    suc_sesion = session.get("sucursal","")
    fecha_iso = request.args.get("fecha", fecha_hoy())
    try:
        fecha_dia = datetime.strptime(fecha_iso, "%Y-%m-%d").strftime("%d-%m-%Y")
    except:
        fecha_dia = fecha_iso

    sucursal = request.args.get("sucursal", suc_sesion)
    if tipo_usuario != "admin":
        sucursal = suc_sesion

    def has_col(cur, table, col):
        cur.execute(f"PRAGMA table_info({table})")
        return any(r[1].lower() == col.lower() for r in cur.fetchall())

    try:
        conn = sqlite3.connect(DB_PATH)
        cur = conn.cursor()

        cur.execute("""
            SELECT IFNULL(SUM(precio),0), COUNT(*)
            FROM ventas
            WHERE fecha LIKE ? AND sucursal=? AND (eliminado IS NULL OR eliminado=0)
        """, (f"{fecha_dia}%", sucursal))
        v_total, v_count = cur.fetchone() or (0,0)

        efectivo = tarjeta = dolar_mxn = mixto = 0.0

        if has_col(cur, "ventas", "tipo_pago"):
            cur.execute("""
                SELECT
                  IFNULL(SUM(CASE WHEN LOWER(IFNULL(tipo_pago,''))='efectivo' THEN precio ELSE 0 END),0),
                  IFNULL(SUM(CASE WHEN LOWER(IFNULL(tipo_pago,''))='tarjeta'  THEN precio ELSE 0 END),0),
                  IFNULL(SUM(CASE WHEN LOWER(IFNULL(tipo_pago,'')) IN ('dolar','dólar','dolares','dólares') THEN precio ELSE 0 END),0),
                  IFNULL(SUM(CASE WHEN LOWER(IFNULL(tipo_pago,''))='mixto'    THEN precio ELSE 0 END),0)
                FROM ventas
                WHERE fecha LIKE ? AND sucursal=? AND (eliminado IS NULL OR eliminado=0)
            """, (f"{fecha_dia}%", sucursal))
            efectivo, tarjeta, dolar_mxn, mixto = cur.fetchone() or (0,0,0,0)
        elif has_col(cur, "ventas_desglose", "tipo_pago"):
            cur.execute("""
                SELECT
                  IFNULL(SUM(CASE WHEN LOWER(tipo_pago)='efectivo' THEN monto ELSE 0 END),0),
                  IFNULL(SUM(CASE WHEN LOWER(tipo_pago)='tarjeta'  THEN monto ELSE 0 END),0),
                  IFNULL(SUM(CASE WHEN LOWER(tipo_pago) IN ('dolar','dólar') THEN monto ELSE 0 END),0)
                FROM ventas_desglose
                WHERE fecha LIKE ? AND sucursal=?
            """, (f"{fecha_dia}%", sucursal))
            efectivo, tarjeta, dolar_mxn = cur.fetchone() or (0,0,0)
            mixto = max(0.0, v_total - (efectivo + tarjeta + dolar_mxn))
        else:
            for col, var in (("efectivo","efectivo"),("tarjeta","tarjeta"),("dolares","dolar_mxn")):
                if has_col(cur, "ventas", col):
                    cur.execute(f"SELECT IFNULL(SUM({col}),0) FROM ventas WHERE fecha LIKE ? AND sucursal=? AND (eliminado IS NULL OR eliminado=0)", (f"{fecha_dia}%", sucursal))
                    val = cur.fetchone()[0] or 0
                    if var=="efectivo": efectivo = val
                    elif var=="tarjeta": tarjeta = val
                    else: dolar_mxn = val
            mixto = max(0.0, v_total - (efectivo + tarjeta + dolar_mxn))

        cur.execute("SELECT IFNULL(SUM(monto),0), COUNT(*) FROM gastos WHERE fecha LIKE ? AND sucursal=?", (f"{fecha_dia}%", sucursal))
        g_total, g_count = cur.fetchone() or (0,0)

        cur.execute("SELECT IFNULL(SUM(monto),0), COUNT(*) FROM anticipos WHERE fecha LIKE ? AND sucursal=?", (f"{fecha_dia}%", sucursal))
        a_total, a_count = cur.fetchone() or (0,0)

        if has_col(cur, "dolares", "monto"):
            cur.execute("SELECT IFNULL(SUM(monto),0), IFNULL(SUM(cantidad),0) FROM dolares WHERE fecha LIKE ? AND sucursal=?", (f"{fecha_dia}%", sucursal))
            d_mxn_total, d_pzas = cur.fetchone() or (0,0)
        else:
            d_mxn_total, d_pzas = 0, 0

        fondo = 0
        if has_col(cur, "fondo_caja", "monto"):
            cur.execute("SELECT monto FROM fondo_caja WHERE fecha LIKE ? AND sucursal=? ORDER BY id DESC LIMIT 1", (f"{fecha_dia}%", sucursal))
            row_f = cur.fetchone()
            fondo = row_f[0] if row_f else 0

        total_notas = 0.0
        if has_col(cur, "notas", "texto"):
            cur.execute("SELECT texto FROM notas WHERE fecha LIKE ? AND sucursal=?", (f"{fecha_dia}%", sucursal))
            for (texto,) in cur.fetchall():
                for x in re.findall(r"\$\s*([\d,]+(?:\.\d{1,2})?)", texto or ""):
                    try: total_notas += float(x.replace(",",""))
                    except: pass

        conn.close()

        return jsonify(
            ok=True,
            fecha=fecha_iso, fecha_dia=fecha_dia, sucursal=sucursal,
            ventas_total=round(v_total,2), ventas_count=int(v_count or 0),
            efectivo=round(efectivo,2), tarjeta=round(tarjeta,2),
            dolar_en_mxn=round(dolar_mxn,2), mixto=round(mixto,2),
            gastos_total=round(g_total,2), gastos_count=int(g_count or 0),
            anticipos_total=round(a_total,2), anticipos_count=int(a_count or 0),
            dolares_mxn=round(d_mxn_total,2), dolares_pzas=int(d_pzas or 0),
            fondo_caja=round(fondo,2), notas_total=round(total_notas,2),
            balance_dia=round((v_total or 0) - (g_total or 0),2)
        )
    except Exception as e:
        print("API resumen ERROR:", e)
        try: conn.close()
        except: pass
        return jsonify(ok=False, msg="server-error")

@app.route("/reporte-excel")
def reporte_excel():
    if "usuario" not in session:
        return redirect("/")

    tipo_usuario = session.get("tipo")
    sucursal = request.args.get("sucursal", session.get("sucursal", ""))
    fecha = request.args.get("fecha", fecha_hoy())  

    
    try:
        fecha = datetime.strptime(fecha, "%Y-%m-%d").strftime("%d-%m-%Y")
    except:
        pass 

    if tipo_usuario != "admin":
        sucursal = session.get("sucursal", "")
        fecha = fecha_hoy()

    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()

    cur.execute("""
        SELECT descripcion, concepto, tipo, referencia, precio, pagado, cambio 
        FROM ventas 
        WHERE fecha LIKE ? AND sucursal = ? AND (eliminado IS NULL OR eliminado = 0)
    """, (f"{fecha}%", sucursal))
    ventas = cur.fetchall()

    cur.execute("SELECT motivo, monto FROM gastos WHERE fecha LIKE ? AND sucursal = ?", (f"{fecha}%", sucursal))
    gastos = cur.fetchall()

    cur.execute("SELECT monto FROM fondo_caja WHERE fecha LIKE ? AND sucursal = ? ORDER BY id DESC LIMIT 1", (f"{fecha}%", sucursal))
    fondo = cur.fetchone()

    cur.execute("SELECT tipo, cantidad, monto FROM dolares WHERE fecha LIKE ? AND sucursal = ?", (f"{fecha}%", sucursal))
    dolares = cur.fetchall()

    cur.execute("SELECT monto FROM anticipos WHERE fecha LIKE ? AND sucursal = ?", (f"{fecha}%", sucursal))
    anticipos = cur.fetchall()

    cur.execute("SELECT texto FROM notas WHERE fecha LIKE ? AND sucursal = ?", (f"{fecha}%", sucursal))
    notas_adicionales = cur.fetchall()

    conn.close()

    plantilla_path = "CORTE.xlsx"
    wb = load_workbook(plantilla_path)
    nueva_hoja = wb.copy_worksheet(wb.active)
    nueva_hoja.title = f"{sucursal}_{fecha}"
    ws = nueva_hoja

    fila = 2
    for venta in ventas:
        ws[f"A{fila}"] = fecha
        ws[f"B{fila}"] = venta[0]
        ws[f"C{fila}"] = venta[2]
        ws[f"D{fila}"] = venta[1]
        ws[f"E{fila}"] = venta[3]
        ws[f"F{fila}"] = ""
        ws[f"G{fila}"] = venta[4]
        fila += 1

    ws["K4"] = sucursal

    if fondo:
        ws["I5"] = fondo[0]

    fila_g = 8
    for g in gastos:
        ws[f"K{fila_g}"] = g[0]
        ws[f"M{fila_g}"] = g[1]
        fila_g += 1
    ws["M41"] = "=SUM(M8:M40)"

    fila_a = 7
    for a in anticipos:
        ws[f"W{fila_a}"] = a[0]
        fila_a += 1
    ws["W41"] = "=SUM(W7:W40)"

    fila_d = 7
    for d in dolares:
        ws[f"R{fila_d}"] = d[0]
        ws[f"S{fila_d}"] = d[1]
        ws[f"T{fila_d}"] = d[2]
        fila_d += 1
    ws["T41"] = "=SUM(T7:T40)"

    fila_nota = 7
    total_notas = 0
    for nota in notas_adicionales:
        texto = nota[0]
        ws[f"U{fila_nota}"] = texto
        match = re.search(r"\$\s?(\d+(?:\.\d{1,2})?)", texto)
        if match:
            try:
                total_notas += float(match.group(1))
            except:
                pass
        fila_nota += 1
    ws["X3"] = total_notas

    if "CORTE DIARIO" in wb.sheetnames:
        wb.remove(wb["CORTE DIARIO"])

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, download_name=f"corte_{sucursal}_{fecha}.xlsx", as_attachment=True)


@app.route("/corte-del-dia")
def corte_del_dia():
    if "usuario" not in session:
        return redirect("/")

    tipo = session.get("tipo", "")
    sucursal_sesion = session.get("sucursal", "")
    tz = ZoneInfo("America/Monterrey")

    if tipo == "admin":
        fecha_qs = request.args.get("fecha", datetime.now(tz).strftime("%Y-%m-%d"))
        try:
            fecha_str = datetime.strptime(fecha_qs, "%Y-%m-%d").strftime("%d-%m-%Y")
        except Exception:
            fecha_str = fecha_qs
        sucursal = request.args.get("sucursal", sucursal_sesion)
    else:
        fecha_str = datetime.now(tz).strftime("%d-%m-%Y")
        sucursal = sucursal_sesion

    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()

    cur.execute("""
        SELECT name FROM sqlite_master
        WHERE type='table' AND name='ventas_desglose'
    """)
    tiene_tabla_desglose = cur.fetchone() is not None

    ventas_efectivo = 0.0
    ventas_tarjeta = 0.0

    if tiene_tabla_desglose:
        cur.execute("""
            SELECT COUNT(*) FROM ventas_desglose
            WHERE fecha LIKE ? AND sucursal = ?
        """, (f"{fecha_str}%", sucursal))
        cnt_desglose = cur.fetchone()[0] or 0
    else:
        cnt_desglose = 0

    if cnt_desglose > 0:
        cur.execute("""
            SELECT
                IFNULL(SUM(COALESCE(efectivo,0)), 0)      AS s_efectivo,
                IFNULL(SUM(COALESCE(tarjeta,0)), 0)       AS s_tarjeta,
                IFNULL(SUM(COALESCE(dolares,0) * COALESCE(tipo_cambio,0)), 0) AS s_dolares_mxn
            FROM ventas_desglose
            WHERE fecha LIKE ? AND sucursal = ?
        """, (f"{fecha_str}%", sucursal))
        s_efec, s_tar, s_dmxn = cur.fetchone()
        ventas_efectivo = (s_efec or 0) + (s_dmxn or 0)
        ventas_tarjeta  = (s_tar or 0)
    else:
        cur.execute("PRAGMA table_info(ventas)")
        cols = {r[1] for r in cur.fetchall()}
        tiene_desglose_en_ventas = {"efectivo", "tarjeta", "dolares", "dolar"}.issubset(cols)

        if tiene_desglose_en_ventas:
            cur.execute("""
                SELECT
                    IFNULL(SUM(COALESCE(efectivo,0)),0),
                    IFNULL(SUM(COALESCE(dolares,0) * COALESCE(dolar,0)),0)
                FROM ventas
                WHERE fecha LIKE ? AND sucursal = ?
            """, (f"{fecha_str}%", sucursal))
            s_efe, s_dol_mxn = cur.fetchone()
            ventas_efectivo = (s_efe or 0) + (s_dol_mxn or 0)

            cur.execute("""
                SELECT IFNULL(SUM(COALESCE(tarjeta,0)),0)
                FROM ventas
                WHERE fecha LIKE ? AND sucursal = ?
            """, (f"{fecha_str}%", sucursal))
            ventas_tarjeta = cur.fetchone()[0] or 0
        else:
            cur.execute("""
                SELECT IFNULL(SUM(precio),0)
                FROM ventas
                WHERE fecha LIKE ? AND sucursal = ?
                  AND (tipo_pago IN ('Efectivo','Dólar') OR metodo_pago IN ('Efectivo','Dólar'))
            """, (f"{fecha_str}%", sucursal))
            ventas_efectivo = cur.fetchone()[0] or 0

            cur.execute("""
                SELECT IFNULL(SUM(precio),0)
                FROM ventas
                WHERE fecha LIKE ? AND sucursal = ?
                  AND (tipo_pago='Tarjeta' OR metodo_pago='Tarjeta')
            """, (f"{fecha_str}%", sucursal))
            ventas_tarjeta = cur.fetchone()[0] or 0

            cur.execute("""
                SELECT IFNULL(SUM(precio),0)
                FROM ventas
                WHERE fecha LIKE ? AND sucursal = ?
                  AND (tipo_pago='Mixto' OR metodo_pago='Mixto')
            """, (f"{fecha_str}%", sucursal))
            mixto_total = cur.fetchone()[0] or 0

            ventas_efectivo += (mixto_total / 2.0)
            ventas_tarjeta  += (mixto_total / 2.0)

    cur.execute("""
        SELECT IFNULL(SUM(monto),0)
        FROM gastos
        WHERE fecha = ? AND sucursal = ?
    """, (fecha_str, sucursal))
    gastos = cur.fetchone()[0] or 0

    cur.execute("""
        SELECT monto FROM fondo_caja
        WHERE fecha = ? AND sucursal = ?
        ORDER BY id DESC LIMIT 1
    """, (fecha_str, sucursal))
    fila_fondo = cur.fetchone()
    fondo_inicial = fila_fondo[0] if fila_fondo else 0

    total_entradas_efectivo = (fondo_inicial or 0) + (ventas_efectivo or 0)
    total_caja = total_entradas_efectivo - (gastos or 0)
    ventas_total_dia = (ventas_efectivo or 0) + (ventas_tarjeta or 0)

    conn.close()

    registrar_log(session["usuario"], tipo, f"Consultó corte del día ({sucursal} - {fecha_str})")

    return render_template(
        "corte-del-dia.html",
        fecha=fecha_str,
        sucursal=sucursal,
        ventas_efectivo=ventas_efectivo,
        ventas_tarjeta=ventas_tarjeta,
        ventas_total_dia=ventas_total_dia,
        gastos=gastos,
        fondo_inicial=fondo_inicial,
        total_entradas_efectivo=total_entradas_efectivo,
        total_caja=total_caja,
        tipo=tipo,
        sucursales=SUCURSALES
    )

if not os.path.exists(DB_PATH):
    init_db()

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5003, debug=True)